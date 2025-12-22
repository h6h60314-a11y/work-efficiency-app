import io
import re
import uuid
import datetime as dt
from typing import Dict, List, Tuple, Optional

import streamlit as st
import pandas as pd

from common_ui import (
    inject_logistics_theme,
    set_page,
    KPI,
    render_kpis,
    bar_topN,
    card_open,
    card_close,
)

from audit_store import sha256_bytes, upload_export_bytes, insert_audit_run

# =========================
# Session Keysï¼ˆç¢ºä¿åŒ¯å‡ºä¸æ¸…ç©º KPIï¼‰
# =========================
RESULT_KEY = "putaway_kpi_result_v1"
AUDIT_SIG_KEY = "putaway_last_audit_sig_v1"


# =========================
# è¦å‰‡ï¼ˆä¾ä¸Šæ¶ v8.9ï¼‰
# =========================
TO_EXCLUDE_KEYWORDS = ["CGS", "JCPL", "QC99", "GREAT0001X", "GX010", "PD99"]
TO_EXCLUDE_PATTERN = re.compile("|".join(re.escape(k) for k in TO_EXCLUDE_KEYWORDS), flags=re.IGNORECASE)

INPUT_USER_CANDIDATES = ["è¨˜éŒ„è¼¸å…¥äºº", "è¨˜éŒ„è¼¸å…¥è€…", "å»ºç«‹äºº", "è¼¸å…¥äºº"]
REV_DT_CANDIDATES = [
    "ä¿®è¨‚æ—¥æœŸ", "ä¿®è¨‚æ™‚é–“", "ä¿®è¨‚æ—¥", "ç•°å‹•æ™‚é–“", "ä¿®æ”¹æ™‚é–“",
    "ä¿®è¨‚æ—¥æœŸæ™‚é–“", "ä¿®è¨‚æ—¥æœŸæ™‚é–“(ç³»çµ±)", "ä¿®è¨‚æ—¥æœŸæ™‚é–“ï¼ˆç³»çµ±ï¼‰",
]

TARGET_EFF = 20
IDLE_MIN_THRESHOLD = 10

AM_START, AM_END = dt.time(7, 0, 0), dt.time(12, 30, 0)
PM_START, PM_END = dt.time(13, 30, 0), dt.time(23, 59, 59)

NAME_MAP = {
    "20200924001": "é»ƒé›…å›", "20210805001": "éƒ­ä¸­åˆ", "20220505002": "é˜®æ–‡é’æ˜",
    "20221221001": "é˜®æ–‡å…¨", "20221222005": "è¬å¿ é¾", "20230119001": "é™¶æ˜¥é’",
    "20240926001": "é™³è‰å¨œ", "20241011002": "æ—é›™æ…§", "20250502001": "å³è©©æ•",
    "20250617001": "é˜®æ–‡è­š", "20250617003": "å–¬å®¶å¯¶", "20250901009": "å¼µå¯¶è±",
    "G01": "0", "20201109003": "å³æŒ¯å‡±", "09963": "é»ƒè¬™å‡±",
    "20240313003": "é˜®æ›°å¿ ", "20201109001": "æ¢å† å¦‚", "10003": "æèŒ‚éŠ“",
    "20200922002": "è‘‰æ¬²å¼˜", "20250923019": "é˜®æ°ç´…æ·±", "9963": "é»ƒè¬™å‡±",
    "11399": "é™³å“²æ²…",
}

BREAK_RULES = [
    (dt.time(20,45,0), dt.time(22,30,0),  0, "é¦–â‰¥20:45 ä¸” æœ«â‰¤22:30 â†’ 0 åˆ†é˜"),
    (dt.time(18,30,0), dt.time(20,30,0),  0, "é¦–â‰¥18:30 ä¸” æœ«â‰¤20:30 â†’ 0 åˆ†é˜"),
    (dt.time(15,30,0), dt.time(18, 0,0),  0, "é¦–â‰¥15:30 ä¸” æœ«â‰¤18:00 â†’ 0 åˆ†é˜"),
    (dt.time(13,30,0), dt.time(15,35,0),  0, "é¦–â‰¥13:30 ä¸” æœ«â‰¤15:35 â†’ 0 åˆ†é˜"),
    (dt.time(20,45,0), dt.time(23, 0,0),  0, "é¦–â‰¥20:45 ä¸” æœ«â‰¤23:00 â†’ 0 åˆ†é˜"),
    (dt.time(20, 0,0), dt.time(22, 0,0), 15, "é¦–â‰¥20:00 ä¸” æœ«â‰¤22:00 â†’ 15 åˆ†é˜"),
    (dt.time(18,30,0), dt.time(22, 0,0), 15, "é¦–â‰¥18:30 ä¸” æœ«â‰¤22:00 â†’ 15 åˆ†é˜"),
    (dt.time(19, 0,0), dt.time(22,30,0), 15, "é¦–â‰¥19:00 ä¸” æœ«â‰¤22:30 â†’ 15 åˆ†é˜"),
    (dt.time(13,30,0), dt.time(18, 0,0), 15, "é¦–â‰¥13:30 ä¸” æœ«â‰¤18:00 â†’ 15 åˆ†é˜"),
    (dt.time(16, 0,0), dt.time(20,40,0), 30, "é¦–â‰¥16:00 ä¸” æœ«â‰¤20:40 â†’ 30 åˆ†é˜"),
    (dt.time(15,30,0), dt.time(20,30,0), 30, "é¦–â‰¥15:30 ä¸” æœ«â‰¤20:30 â†’ 30 åˆ†é˜"),
    (dt.time(17, 0,0), dt.time(22,30,0), 45, "é¦–â‰¥17:00 ä¸” æœ«â‰¤22:30 â†’ 45 åˆ†é˜"),
    (dt.time(15,45,0), dt.time(22,30,0), 45, "é¦–â‰¥15:45 ä¸” æœ«â‰¤22:30 â†’ 45 åˆ†é˜"),
    (dt.time(13,30,0), dt.time(20,29,0), 45, "é¦–â‰¥13:30 ä¸” æœ«â‰¤20:29 â†’ 45 åˆ†é˜"),
    (dt.time(13,30,0), dt.time(23, 0,0), 60, "é¦–â‰¥13:30 ä¸” æœ«â‰¤23:00 â†’ 60 åˆ†é˜"),
    (dt.time(11, 0,0), dt.time(17, 0,0), 75, "é¦–â‰¥11:00 ä¸” æœ«â‰¤17:00 â†’ 75 åˆ†é˜"),
    (dt.time( 8, 0,0), dt.time(17, 0,0), 90, "é¦–â‰¥08:00 ä¸” æœ«â‰¤17:00 â†’ 90 åˆ†é˜"),
    (dt.time(10,50,0), dt.time(23, 0,0),120, "é¦–â‰¥10:50 ä¸” æœ«â‰¤23:00 â†’ 120 åˆ†é˜"),
    (dt.time( 8, 0,0), dt.time(23, 0,0),135, "é¦–â‰¥08:00 ä¸” æœ«â‰¤23:00 â†’ 135 åˆ†é˜"),
]

EXCLUDE_IDLE_RANGES = [
    (dt.time(10, 0, 0), dt.time(10, 15, 0)),
    (dt.time(12,30, 0), dt.time(13, 30, 0)),
    (dt.time(15,30, 0), dt.time(15, 45, 0)),
    (dt.time(18, 0, 0), dt.time(18, 30, 0)),
    (dt.time(20,30, 0), dt.time(20, 45, 0)),
]


# =========================
# è®€æª”ï¼šxlsx/xlsm/xls/csv
# =========================
def read_excel_any_quiet_bytes(name: str, content: bytes) -> Dict[str, pd.DataFrame]:
    ext = (name.split(".")[-1] or "").lower()

    if ext in ("xlsx", "xlsm"):
        xl = pd.ExcelFile(io.BytesIO(content), engine="openpyxl")
        return {sn: pd.read_excel(xl, sheet_name=sn) for sn in xl.sheet_names}

    # èˆŠç‰ˆ .xlsï¼šéœ€è¦ requirements.txt å®‰è£ xlrd==2.0.1
    if ext == "xls":
        xl = pd.ExcelFile(io.BytesIO(content), engine="xlrd")
        return {sn: pd.read_excel(xl, sheet_name=sn) for sn in xl.sheet_names}

    if ext == "csv":
        for enc in ("utf-8-sig", "cp950", "big5"):
            try:
                return {"CSV": pd.read_csv(io.BytesIO(content), encoding=enc)}
            except Exception:
                continue
        raise Exception("CSV è®€å–å¤±æ•—ï¼ˆè«‹ç¢ºèªç·¨ç¢¼ï¼‰ã€‚")

    raise Exception("ç›®å‰åƒ…æ”¯æ´ .xlsx/.xlsm/.xls/.csv")


# =========================
# å·¥å…·
# =========================
def _strip_cols(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.columns = [str(c).strip() for c in df.columns]
    return df

def find_first_column(df: pd.DataFrame, candidates: List[str]) -> Optional[str]:
    cols = [str(c).strip() for c in df.columns]
    s = set(cols)
    for name in candidates:
        if name in s:
            return name
    norm_map = {re.sub(r"[ï¼ˆï¼‰\(\)\s]", "", c): c for c in cols}
    for name in candidates:
        key = re.sub(r"[ï¼ˆï¼‰\(\)\s]", "", name)
        if key in norm_map:
            return norm_map[key]
    return None

def normalize_to_qc(series: pd.Series) -> pd.Series:
    s = series.astype(str).str.strip().str.upper()
    return s.eq("QC")

def to_not_excluded_mask(series: pd.Series) -> pd.Series:
    s = series.astype(str).str.strip()
    return ~s.str.contains(TO_EXCLUDE_PATTERN, na=False)

def prepare_filtered_df(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame()
    df = _strip_cols(df)
    if "ç”±" not in df.columns or "åˆ°" not in df.columns:
        return pd.DataFrame()
    return df[normalize_to_qc(df["ç”±"]) & to_not_excluded_mask(df["åˆ°"])].copy()

def break_minutes_for_span(first_dt: pd.Timestamp, last_dt: pd.Timestamp) -> Tuple[int, str]:
    if pd.isna(first_dt) or pd.isna(last_dt):
        return 0, "ç„¡æ™‚é–“è³‡æ–™"
    stt, edt = first_dt.time(), last_dt.time()
    for st_ge, ed_le, mins, tag in BREAK_RULES:
        if (stt >= st_ge) and (edt <= ed_le):
            return int(mins), str(tag)
    return 0, "æœªå‘½ä¸­è¦å‰‡"

def _subtract_exclusions(s_dt: pd.Timestamp, e_dt: pd.Timestamp, exclude_ranges):
    if s_dt >= e_dt or not exclude_ranges:
        return [(s_dt, e_dt)]
    segments = [(s_dt, e_dt)]
    for ex_s_t, ex_e_t in exclude_ranges:
        ex_s = pd.Timestamp.combine(s_dt.date(), ex_s_t)
        ex_e = pd.Timestamp.combine(s_dt.date(), ex_e_t)
        new_segments = []
        for a, b in segments:
            if b <= ex_s or a >= ex_e:
                new_segments.append((a, b))
            else:
                if a < ex_s:
                    new_segments.append((a, ex_s))
                if b > ex_e:
                    new_segments.append((ex_e, b))
        segments = [(x, y) for (x, y) in new_segments if x < y]
    return segments

def _compute_idle(series_dt: pd.Series, min_minutes=IDLE_MIN_THRESHOLD, exclude_ranges=EXCLUDE_IDLE_RANGES) -> Tuple[int, str]:
    if series_dt.size < 2:
        return 0, ""
    s = series_dt.sort_values()
    total_min, ranges_txt = 0, []
    prev = s.iloc[0]
    for cur in s.iloc[1:]:
        if cur <= prev:
            prev = cur
            continue
        for a, b in _subtract_exclusions(prev, cur, exclude_ranges or []):
            gap_min = int(round((b - a).total_seconds() / 60.0))
            if gap_min >= min_minutes:
                total_min += gap_min
                ranges_txt.append(f"{a.time()} ~ {b.time()}")
        prev = cur
    return int(total_min), "ï¼›".join(ranges_txt)

def _span_metrics(series_dt: pd.Series):
    if series_dt.empty:
        return pd.NaT, pd.NaT, 0
    return series_dt.min(), series_dt.max(), int(series_dt.size)

def _eff(n, m):
    return round((n / m * 60.0), 2) if m and m > 0 else 0.0

def compute_am_pm_for_group(g: pd.DataFrame) -> pd.Series:
    times = g["__dt__"]

    # AMï¼ˆä¸æ‰£ä¼‘ï¼‰
    t_am = times[times.dt.time.between(AM_START, AM_END)]
    am_first, am_last, am_cnt = _span_metrics(t_am)
    am_mins = int(round(((am_last - am_first).total_seconds() / 60.0))) if am_cnt > 0 else 0
    am_eff = _eff(am_cnt, am_mins)
    am_idle_min, am_idle_ranges = _compute_idle(t_am)

    # PMï¼ˆæ‰£ä¼‘ï¼‰
    t_pm = times[times.dt.time.between(PM_START, PM_END)]
    pm_first, pm_last, pm_cnt = _span_metrics(t_pm)
    if pm_cnt > 0:
        pm_break, pm_rule = break_minutes_for_span(pm_first, pm_last)
        raw_pm_mins = (pm_last - pm_first).total_seconds() / 60.0
        pm_mins = max(int(round(raw_pm_mins - pm_break)), 0)
    else:
        pm_break, pm_rule, pm_mins = 0, "ç„¡æ™‚é–“è³‡æ–™", 0
    pm_eff = _eff(pm_cnt, pm_mins)
    pm_idle_min, pm_idle_ranges = _compute_idle(t_pm)

    # Wholeï¼ˆæ‰£ä¼‘ï¼‰
    whole_first, whole_last, day_cnt = _span_metrics(times)
    if day_cnt > 0:
        whole_break, br_tag_whole = break_minutes_for_span(whole_first, whole_last)
        raw_whole_mins = (whole_last - whole_first).total_seconds() / 60.0
        whole_mins = max(int(round(raw_whole_mins - whole_break)), 0)
    else:
        whole_break, br_tag_whole, whole_mins = 0, "ç„¡æ™‚é–“è³‡æ–™", 0
    whole_eff = _eff(day_cnt, whole_mins)

    return pd.Series({
        "ç¬¬ä¸€ç­†æ™‚é–“": whole_first, "æœ€å¾Œä¸€ç­†æ™‚é–“": whole_last, "ç•¶æ—¥ç­†æ•¸": int(day_cnt),
        "ä¼‘æ¯åˆ†é˜_æ•´é«”": int(whole_break), "å‘½ä¸­è¦å‰‡": br_tag_whole,
        "ç•¶æ—¥å·¥æ™‚_åˆ†é˜_æ‰£ä¼‘": int(whole_mins), "æ•ˆç‡_ä»¶æ¯å°æ™‚": whole_eff,

        "ä¸Šåˆ_ç¬¬ä¸€ç­†": am_first, "ä¸Šåˆ_æœ€å¾Œä¸€ç­†": am_last, "ä¸Šåˆ_ç­†æ•¸": int(am_cnt),
        "ä¸Šåˆ_å·¥æ™‚_åˆ†é˜": int(am_mins), "ä¸Šåˆ_æ•ˆç‡_ä»¶æ¯å°æ™‚": am_eff,
        "ä¸Šåˆ_ç©ºçª—åˆ†é˜": int(am_idle_min), "ä¸Šåˆ_ç©ºçª—æ™‚æ®µ": am_idle_ranges,

        "ä¸‹åˆ_ç¬¬ä¸€ç­†": pm_first, "ä¸‹åˆ_æœ€å¾Œä¸€ç­†": pm_last, "ä¸‹åˆ_ç­†æ•¸": int(pm_cnt),
        "ä¸‹åˆ_ä¼‘æ¯åˆ†é˜": int(pm_break), "ä¸‹åˆ_å‘½ä¸­è¦å‰‡": pm_rule,
        "ä¸‹åˆ_å·¥æ™‚_åˆ†é˜_æ‰£ä¼‘": int(pm_mins), "ä¸‹åˆ_æ•ˆç‡_ä»¶æ¯å°æ™‚": pm_eff,
        "ä¸‹åˆ_ç©ºçª—åˆ†é˜_æ‰£ä¼‘": int(pm_idle_min), "ä¸‹åˆ_ç©ºçª—æ™‚æ®µ": pm_idle_ranges,
    })


# =========================
# åŒ¯å‡º Excelï¼ˆbytesï¼‰
# =========================
def autosize_columns(ws, df: pd.DataFrame):
    from openpyxl.utils import get_column_letter
    cols = list(df.columns) if df is not None else []
    for i, col in enumerate(cols, start=1):
        if df is not None and not df.empty:
            sample = [len(str(x)) for x in df[col].head(800).tolist()]
            max_len = max([len(str(col))] + sample)
        else:
            max_len = max(len(str(col)), 8)
        ws.column_dimensions[get_column_letter(i)].width = min(max_len + 2, 60)

def shade_rows_by_efficiency(ws, header_name="æ•ˆç‡_ä»¶æ¯å°æ™‚", green="C6EFCE", red="FFC7CE"):
    from openpyxl.styles import PatternFill
    eff_col = None
    for c in range(1, ws.max_column + 1):
        if str(ws.cell(row=1, column=c).value).strip() == header_name:
            eff_col = c
            break
    if eff_col is None:
        return
    green_fill = PatternFill(start_color=green, end_color=green, fill_type="solid")
    red_fill = PatternFill(start_color=red, end_color=red, fill_type="solid")
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=eff_col).value
        try:
            val = float(v) if v is not None and str(v).strip() != "" else None
        except Exception:
            val = None
        if val is None:
            continue
        fill = green_fill if val >= TARGET_EFF else red_fill
        for c in range(1, ws.max_column + 1):
            ws.cell(row=r, column=c).fill = fill

def build_excel_bytes(user_col: str, summary_out: pd.DataFrame, daily: pd.DataFrame, detail_long: pd.DataFrame) -> bytes:
    out = io.BytesIO()
    with pd.ExcelWriter(out, engine="openpyxl", datetime_format="yyyy-mm-dd hh:mm:ss", date_format="yyyy-mm-dd") as writer:
        sum_cols = [
            user_col, "å°æ‡‰å§“å", "ç·æ—¥æ•¸",
            "ç¸½ç­†æ•¸", "ç¸½å·¥æ™‚_åˆ†é˜_æ‰£ä¼‘", "æ•ˆç‡_ä»¶æ¯å°æ™‚",
            "ä¸Šåˆç­†æ•¸", "ä¸Šåˆå·¥æ™‚_åˆ†é˜", "ä¸Šåˆæ•ˆç‡_ä»¶æ¯å°æ™‚",
            "ä¸‹åˆç­†æ•¸", "ä¸‹åˆå·¥æ™‚_åˆ†é˜_æ‰£ä¼‘", "ä¸‹åˆæ•ˆç‡_ä»¶æ¯å°æ™‚",
        ]
        summary_out[sum_cols].to_excel(writer, index=False, sheet_name="å½™ç¸½")
        ws_sum = writer.sheets["å½™ç¸½"]
        autosize_columns(ws_sum, summary_out[sum_cols])
        shade_rows_by_efficiency(ws_sum, "æ•ˆç‡_ä»¶æ¯å°æ™‚")

        det_cols = [
            user_col, "å°æ‡‰å§“å", "æ—¥æœŸ",
            "ç¬¬ä¸€ç­†æ™‚é–“", "æœ€å¾Œä¸€ç­†æ™‚é–“", "ç•¶æ—¥ç­†æ•¸",
            "ä¼‘æ¯åˆ†é˜_æ•´é«”", "ç•¶æ—¥å·¥æ™‚_åˆ†é˜_æ‰£ä¼‘", "æ•ˆç‡_ä»¶æ¯å°æ™‚",
            "ä¸Šåˆ_ç¬¬ä¸€ç­†", "ä¸Šåˆ_æœ€å¾Œä¸€ç­†", "ä¸Šåˆ_ç­†æ•¸", "ä¸Šåˆ_å·¥æ™‚_åˆ†é˜", "ä¸Šåˆ_æ•ˆç‡_ä»¶æ¯å°æ™‚",
            "ä¸Šåˆ_ç©ºçª—åˆ†é˜", "ä¸Šåˆ_ç©ºçª—æ™‚æ®µ",
            "ä¸‹åˆ_ç¬¬ä¸€ç­†", "ä¸‹åˆ_æœ€å¾Œä¸€ç­†", "ä¸‹åˆ_ç­†æ•¸", "ä¸‹åˆ_ä¼‘æ¯åˆ†é˜",
            "ä¸‹åˆ_å·¥æ™‚_åˆ†é˜_æ‰£ä¼‘", "ä¸‹åˆ_æ•ˆç‡_ä»¶æ¯å°æ™‚",
            "ä¸‹åˆ_ç©ºçª—åˆ†é˜_æ‰£ä¼‘", "ä¸‹åˆ_ç©ºçª—æ™‚æ®µ",
        ]
        daily.sort_values([user_col, "æ—¥æœŸ", "ç¬¬ä¸€ç­†æ™‚é–“"])[det_cols].to_excel(writer, index=False, sheet_name="æ˜ç´°")
        ws_det = writer.sheets["æ˜ç´°"]
        autosize_columns(ws_det, daily[det_cols])
        shade_rows_by_efficiency(ws_det, "æ•ˆç‡_ä»¶æ¯å°æ™‚")

        if detail_long is not None and not detail_long.empty:
            long_cols = [
                user_col, "å°æ‡‰å§“å", "æ—¥æœŸ", "æ™‚æ®µ",
                "ç¬¬ä¸€ç­†æ™‚é–“", "æœ€å¾Œä¸€ç­†æ™‚é–“",
                "ç­†æ•¸", "å·¥æ™‚_åˆ†é˜", "ä¼‘æ¯åˆ†é˜",
                "ç©ºçª—åˆ†é˜", "ç©ºçª—æ™‚æ®µ",
                "æ•ˆç‡_ä»¶æ¯å°æ™‚", "å‘½ä¸­è¦å‰‡",
            ]
            detail_long[long_cols].to_excel(writer, index=False, sheet_name="æ˜ç´°_æ™‚æ®µ")
            ws_long = writer.sheets["æ˜ç´°_æ™‚æ®µ"]
            autosize_columns(ws_long, detail_long[long_cols])
            shade_rows_by_efficiency(ws_long, "æ•ˆç‡_ä»¶æ¯å°æ™‚")

        # ä¼‘æ¯è¦å‰‡
        rules_rows = []
        for i, (st_ge, ed_le, mins, tag) in enumerate(BREAK_RULES, start=1):
            rules_rows.append({
                "å„ªå…ˆåº": i,
                "é¦–æ™‚é–“æ¢ä»¶(>=)": st_ge.strftime("%H:%M:%S"),
                "æœ«æ™‚é–“æ¢ä»¶(<=)": ed_le.strftime("%H:%M:%S"),
                "ä¼‘æ¯åˆ†é˜": int(mins),
                "è¦å‰‡èªªæ˜": str(tag),
            })
        rules_df = pd.DataFrame(rules_rows, columns=["å„ªå…ˆåº", "é¦–æ™‚é–“æ¢ä»¶(>=)", "æœ«æ™‚é–“æ¢ä»¶(<=)", "ä¼‘æ¯åˆ†é˜", "è¦å‰‡èªªæ˜"])
        rules_df.to_excel(writer, index=False, sheet_name="ä¼‘æ¯è¦å‰‡")
        autosize_columns(writer.sheets["ä¼‘æ¯è¦å‰‡"], rules_df)

    return out.getvalue()


# =========================
# ä¸»æµç¨‹ï¼šè¨ˆç®— + å­˜ session
# =========================
def compute_and_store(uploaded_name: str, uploaded_bytes: bytes, operator: str, top_n: int):
    sheets = read_excel_any_quiet_bytes(uploaded_name, uploaded_bytes)

    kept_all = []
    for sn, df in sheets.items():
        k = prepare_filtered_df(df)
        if not k.empty:
            k["__sheet__"] = sn
            kept_all.append(k)
    if not kept_all:
        raise Exception("ç„¡ç¬¦åˆè³‡æ–™ï¼ˆå¯èƒ½ç¼ºã€ç”±/åˆ°ã€æ¬„æˆ–éæ¿¾å¾Œç‚ºç©ºï¼‰ã€‚")

    data = pd.concat(kept_all, ignore_index=True)

    user_col = find_first_column(data, INPUT_USER_CANDIDATES)
    revdt_col = find_first_column(data, REV_DT_CANDIDATES)
    if user_col is None:
        raise Exception("æ‰¾ä¸åˆ°ã€è¨˜éŒ„è¼¸å…¥äººã€æ¬„ä½ï¼ˆå€™é¸ï¼šè¨˜éŒ„è¼¸å…¥äºº/è¨˜éŒ„è¼¸å…¥è€…/å»ºç«‹äºº/è¼¸å…¥äººï¼‰ã€‚")
    if revdt_col is None:
        raise Exception("æ‰¾ä¸åˆ°ã€ä¿®è¨‚æ—¥æœŸ/æ™‚é–“ã€æ¬„ä½ï¼ˆå€™é¸ï¼šä¿®è¨‚æ—¥æœŸ/ä¿®è¨‚æ™‚é–“/ç•°å‹•æ™‚é–“/ä¿®æ”¹æ™‚é–“â€¦ï¼‰ã€‚")

    data["__dt__"] = pd.to_datetime(data[revdt_col], errors="coerce")
    data["__code__"] = data[user_col].astype(str).str.strip()
    data["å°æ‡‰å§“å"] = data["__code__"].map(NAME_MAP).fillna("")
    dt_data = data.dropna(subset=["__dt__"]).copy()
    if dt_data.empty:
        raise Exception("è³‡æ–™æ²’æœ‰å¯ç”¨çš„ä¿®è¨‚æ—¥æœŸæ™‚é–“ï¼Œç„¡æ³•è¨ˆç®—ã€‚")

    dt_data["æ—¥æœŸ"] = dt_data["__dt__"].dt.date

    daily = (
        dt_data.groupby([user_col, "å°æ‡‰å§“å", "æ—¥æœŸ"], dropna=False)
        .apply(compute_am_pm_for_group)
        .reset_index()
    )

    summary = (
        daily.groupby([user_col, "å°æ‡‰å§“å"], dropna=False, as_index=False)
        .agg(
            ç·æ—¥æ•¸=("æ—¥æœŸ", "nunique"),
            ç¸½ç­†æ•¸=("ç•¶æ—¥ç­†æ•¸", "sum"),
            ç¸½å·¥æ™‚_åˆ†é˜_æ‰£ä¼‘=("ç•¶æ—¥å·¥æ™‚_åˆ†é˜_æ‰£ä¼‘", "sum"),
            ä¸Šåˆç­†æ•¸=("ä¸Šåˆ_ç­†æ•¸", "sum"),
            ä¸Šåˆå·¥æ™‚_åˆ†é˜=("ä¸Šåˆ_å·¥æ™‚_åˆ†é˜", "sum"),
            ä¸‹åˆç­†æ•¸=("ä¸‹åˆ_ç­†æ•¸", "sum"),
            ä¸‹åˆå·¥æ™‚_åˆ†é˜_æ‰£ä¼‘=("ä¸‹åˆ_å·¥æ™‚_åˆ†é˜_æ‰£ä¼‘", "sum"),
        )
    )
    summary["ä¸Šåˆæ•ˆç‡_ä»¶æ¯å°æ™‚"] = summary.apply(lambda r: _eff(r["ä¸Šåˆç­†æ•¸"], r["ä¸Šåˆå·¥æ™‚_åˆ†é˜"]), axis=1)
    summary["ä¸‹åˆæ•ˆç‡_ä»¶æ¯å°æ™‚"] = summary.apply(lambda r: _eff(r["ä¸‹åˆç­†æ•¸"], r["ä¸‹åˆå·¥æ™‚_åˆ†é˜_æ‰£ä¼‘"]), axis=1)
    summary["ç¸½å·¥æ™‚_åˆ†é˜_æ‰£ä¼‘"] = summary["ä¸Šåˆå·¥æ™‚_åˆ†é˜"].fillna(0).astype(int) + summary["ä¸‹åˆå·¥æ™‚_åˆ†é˜_æ‰£ä¼‘"].fillna(0).astype(int)
    summary["æ•ˆç‡_ä»¶æ¯å°æ™‚"] = summary.apply(lambda r: _eff(r["ç¸½ç­†æ•¸"], r["ç¸½å·¥æ™‚_åˆ†é˜_æ‰£ä¼‘"]), axis=1)

    for c in ["ç¸½ç­†æ•¸", "ç¸½å·¥æ™‚_åˆ†é˜_æ‰£ä¼‘", "ä¸Šåˆç­†æ•¸", "ä¸Šåˆå·¥æ™‚_åˆ†é˜", "ä¸‹åˆç­†æ•¸", "ä¸‹åˆå·¥æ™‚_åˆ†é˜_æ‰£ä¼‘"]:
        summary[c] = summary[c].fillna(0).astype(int)

    total_people = int(summary[user_col].nunique())
    total_met = int((summary["æ•ˆç‡_ä»¶æ¯å°æ™‚"] >= TARGET_EFF).sum())
    total_rate = (total_met / total_people) if total_people > 0 else 0.0

    pm_met = int((summary["ä¸‹åˆæ•ˆç‡_ä»¶æ¯å°æ™‚"] >= TARGET_EFF).sum())
    pm_total = int(summary[user_col].nunique())
    pm_rate = (pm_met / pm_total) if pm_total > 0 else 0.0

    total_row = {
        user_col: "æ•´é«”åˆè¨ˆ", "å°æ‡‰å§“å": "",
        "ç·æ—¥æ•¸": int(summary["ç·æ—¥æ•¸"].sum()),
        "ç¸½ç­†æ•¸": int(summary["ç¸½ç­†æ•¸"].sum()),
        "ç¸½å·¥æ™‚_åˆ†é˜_æ‰£ä¼‘": int(summary["ç¸½å·¥æ™‚_åˆ†é˜_æ‰£ä¼‘"].sum()),
        "ä¸Šåˆç­†æ•¸": int(summary["ä¸Šåˆç­†æ•¸"].sum()),
        "ä¸Šåˆå·¥æ™‚_åˆ†é˜": int(summary["ä¸Šåˆå·¥æ™‚_åˆ†é˜"].sum()),
        "ä¸‹åˆç­†æ•¸": int(summary["ä¸‹åˆç­†æ•¸"].sum()),
        "ä¸‹åˆå·¥æ™‚_åˆ†é˜_æ‰£ä¼‘": int(summary["ä¸‹åˆå·¥æ™‚_åˆ†é˜_æ‰£ä¼‘"].sum()),
        "æ•ˆç‡_ä»¶æ¯å°æ™‚": _eff(int(summary["ç¸½ç­†æ•¸"].sum()), int(summary["ç¸½å·¥æ™‚_åˆ†é˜_æ‰£ä¼‘"].sum())),
        "ä¸Šåˆæ•ˆç‡_ä»¶æ¯å°æ™‚": _eff(int(summary["ä¸Šåˆç­†æ•¸"].sum()), int(summary["ä¸Šåˆå·¥æ™‚_åˆ†é˜"].sum())),
        "ä¸‹åˆæ•ˆç‡_ä»¶æ¯å°æ™‚": _eff(int(summary["ä¸‹åˆç­†æ•¸"].sum()), int(summary["ä¸‹åˆå·¥æ™‚_åˆ†é˜_æ‰£ä¼‘"].sum())),
    }
    summary_out = pd.concat([summary, pd.DataFrame([total_row])], ignore_index=True)

    # æ˜ç´°_æ™‚æ®µï¼ˆé•·è¡¨ï¼‰
    long_rows = []
    for _, r in daily.iterrows():
        if r["ä¸Šåˆ_ç­†æ•¸"] > 0:
            long_rows.append({
                user_col: r[user_col], "å°æ‡‰å§“å": r["å°æ‡‰å§“å"], "æ—¥æœŸ": r["æ—¥æœŸ"], "æ™‚æ®µ": "ä¸Šåˆ",
                "ç¬¬ä¸€ç­†æ™‚é–“": r["ä¸Šåˆ_ç¬¬ä¸€ç­†"], "æœ€å¾Œä¸€ç­†æ™‚é–“": r["ä¸Šåˆ_æœ€å¾Œä¸€ç­†"],
                "ç­†æ•¸": int(r["ä¸Šåˆ_ç­†æ•¸"]),
                "å·¥æ™‚_åˆ†é˜": int(r["ä¸Šåˆ_å·¥æ™‚_åˆ†é˜"]),
                "ä¼‘æ¯åˆ†é˜": 0,
                "ç©ºçª—åˆ†é˜": int(r["ä¸Šåˆ_ç©ºçª—åˆ†é˜"]),
                "ç©ºçª—æ™‚æ®µ": r["ä¸Šåˆ_ç©ºçª—æ™‚æ®µ"],
                "æ•ˆç‡_ä»¶æ¯å°æ™‚": float(r["ä¸Šåˆ_æ•ˆç‡_ä»¶æ¯å°æ™‚"]),
                "å‘½ä¸­è¦å‰‡": "ä¸Šåˆä¸æ‰£ä¼‘",
            })
        if r["ä¸‹åˆ_ç­†æ•¸"] > 0:
            long_rows.append({
                user_col: r[user_col], "å°æ‡‰å§“å": r["å°æ‡‰å§“å"], "æ—¥æœŸ": r["æ—¥æœŸ"], "æ™‚æ®µ": "ä¸‹åˆ",
                "ç¬¬ä¸€ç­†æ™‚é–“": r["ä¸‹åˆ_ç¬¬ä¸€ç­†"], "æœ€å¾Œä¸€ç­†æ™‚é–“": r["ä¸‹åˆ_æœ€å¾Œä¸€ç­†"],
                "ç­†æ•¸": int(r["ä¸‹åˆ_ç­†æ•¸"]),
                "å·¥æ™‚_åˆ†é˜": int(r["ä¸‹åˆ_å·¥æ™‚_åˆ†é˜_æ‰£ä¼‘"]),
                "ä¼‘æ¯åˆ†é˜": int(r["ä¸‹åˆ_ä¼‘æ¯åˆ†é˜"]),
                "ç©ºçª—åˆ†é˜": int(r["ä¸‹åˆ_ç©ºçª—åˆ†é˜_æ‰£ä¼‘"]),
                "ç©ºçª—æ™‚æ®µ": r["ä¸‹åˆ_ç©ºçª—æ™‚æ®µ"],
                "æ•ˆç‡_ä»¶æ¯å°æ™‚": float(r["ä¸‹åˆ_æ•ˆç‡_ä»¶æ¯å°æ™‚"]),
                "å‘½ä¸­è¦å‰‡": str(r["ä¸‹åˆ_å‘½ä¸­è¦å‰‡"]),
            })
    detail_long = pd.DataFrame(long_rows)
    if not detail_long.empty:
        detail_long = detail_long.sort_values([user_col, "æ—¥æœŸ", "æ™‚æ®µ", "ç¬¬ä¸€ç­†æ™‚é–“"])

    xlsx_bytes = build_excel_bytes(user_col, summary_out, daily, detail_long)

    # å­˜ sessionï¼ˆKPI/åœ–è¡¨/åŒ¯å‡ºéƒ½å¾é€™è£¡è®€ï¼ŒåŒ¯å‡ºä¸æœƒæ¸…ç©ºï¼‰
    st.session_state[RESULT_KEY] = {
        "user_col": user_col,
        "summary": summary,
        "summary_out": summary_out,
        "daily": daily,
        "detail_long": detail_long,
        "xlsx_bytes": xlsx_bytes,
        "kpi": {
            "total_people": total_people,
            "total_met": total_met,
            "total_rate": total_rate,
            "pm_total": pm_total,
            "pm_met": pm_met,
            "pm_rate": pm_rate,
        },
        "meta": {
            "operator": operator or None,
            "top_n": int(top_n),
            "source_filename": uploaded_name,
            "source_sha256": sha256_bytes(uploaded_bytes),
        },
    }


def try_audit_persist():
    """
    æŠŠæœ¬æ¬¡çµæœç•™å­˜åˆ° Supabaseï¼ˆDB + Storageï¼‰ï¼Œä¸¦é¿å…åŒä¸€ä»½çµæœåè¦†å¯«å…¥ã€‚
    """
    result = st.session_state.get(RESULT_KEY)
    if not result:
        return

    meta = result["meta"]
    # ç”¨ (ä¾†æºhash + top_n + operator) ç•¶ç°½ç« é¿å…é‡è¤‡å¯«å…¥
    sig = f"{meta['source_sha256']}|top_n={meta['top_n']}|op={meta.get('operator')}"
    if st.session_state.get(AUDIT_SIG_KEY) == sig:
        return  # å·²å¯«é

    xlsx_bytes = result["xlsx_bytes"]
    kpi = result["kpi"]

    export_path = upload_export_bytes(
        content=xlsx_bytes,
        object_path=f"putaway_runs/{dt.datetime.now():%Y%m%d}/{uuid.uuid4().hex}_putaway.xlsx",
    )
    payload = {
        "app_name": "ä¸Šæ¶ç”¢èƒ½åˆ†æï¼ˆPutaway KPIï¼‰",
        "operator": meta.get("operator"),
        "source_filename": meta["source_filename"],
        "source_sha256": meta["source_sha256"],
        "params": {
            "top_n": meta["top_n"],
            "target_eff": TARGET_EFF,
            "filter": "ç”±=QC ä¸” åˆ°ä¸å«é—œéµå­—",
            "am_range": "07:00-12:30",
            "pm_range": "13:30-23:59:59",
            "idle_min_threshold": IDLE_MIN_THRESHOLD,
            "idle_exclude_ranges": [(a.strftime("%H:%M"), b.strftime("%H:%M")) for a, b in EXCLUDE_IDLE_RANGES],
        },
        "kpi_am": {"people": int(kpi["total_people"]), "pass_rate": float(kpi["total_rate"])},
        "kpi_pm": {"people": int(kpi["pm_total"]), "pass_rate": float(kpi["pm_rate"])},
        "export_object_path": export_path,
    }
    row = insert_audit_run(payload)
    st.session_state[AUDIT_SIG_KEY] = sig
    return row


# =========================
# Streamlit Page
# =========================
def main():
    inject_logistics_theme()
    set_page("ä¸Šæ¶ç”¢èƒ½åˆ†æï¼ˆPutaway KPIï¼‰", icon="ğŸ“¦")
    st.caption("ç¸½ä¸Šçµ„ï¼ˆä¸Šæ¶ï¼‰ï½œAM/PMï½œç©ºçª—æ‰£é™¤å›ºå®šå¸¶ï½œä¸‹åˆæ‰£ä¼‘ï½œåŒ¯å‡ºä¸€éµä¸‹è¼‰ï¼ˆä¸æ¸…ç©º KPIï¼‰")

    with st.sidebar:
        st.header("âš™ï¸ è¨­å®š")
        operator = st.text_input("åˆ†æåŸ·è¡Œäººï¼ˆOperatorï¼‰")
        top_n = st.number_input("æ•ˆç‡æ’è¡Œé¡¯ç¤ºäººæ•¸ï¼ˆTop Nï¼‰", 10, 100, 30, step=5)
        st.caption("ä¸Šå‚³ .xls éœ€ requirements.txt åŠ ï¼šxlrd==2.0.1")

        if st.button("ğŸ§¹ æ¸…é™¤æœ¬é çµæœ", use_container_width=True):
            st.session_state.pop(RESULT_KEY, None)
            st.session_state.pop(AUDIT_SIG_KEY, None)
            st.rerun()

    # ä¸Šå‚³å€
    card_open("ğŸ“¤ ä¸Šå‚³ä½œæ¥­åŸå§‹è³‡æ–™ï¼ˆä¸Šæ¶ï¼‰")
    uploaded = st.file_uploader(
        "ä¸Šå‚³ Excel / CSV",
        type=["xlsx", "xlsm", "xls", "csv"],
        label_visibility="collapsed",
    )
    run = st.button("ğŸš€ ç”¢å‡º KPI", type="primary", disabled=uploaded is None)
    card_close()

    # æŒ‰ä¸‹æ‰è¨ˆç®—ï¼›æ²’æŒ‰ä¹Ÿä¸ returnï¼ˆè®“ä¸Šæ¬¡ KPI ä¿ç•™ï¼‰
    if run:
        try:
            with st.spinner("è¨ˆç®—ä¸­ï¼Œè«‹ç¨å€™..."):
                compute_and_store(
                    uploaded_name=uploaded.name,
                    uploaded_bytes=uploaded.getvalue(),
                    operator=operator,
                    top_n=int(top_n),
                )
            st.success("âœ… å·²å®Œæˆ KPI è¨ˆç®—")
        except Exception as e:
            st.error("âŒ è¨ˆç®—å¤±æ•—")
            st.code(repr(e))
            return

    # é¡¯ç¤ºçµæœï¼ˆå¾ sessionï¼‰
    result = st.session_state.get(RESULT_KEY)
    if not result:
        st.info("è«‹å…ˆä¸Šå‚³æª”æ¡ˆä¸¦æŒ‰ã€ç”¢å‡º KPIã€ã€‚")
        return

    user_col = result["user_col"]
    summary = result["summary"]
    xlsx_bytes = result["xlsx_bytes"]
    kpi = result["kpi"]
    meta = result["meta"]

    # å·¦å³ AM/PM
    col_l, col_r = st.columns(2)

    with col_l:
        card_open("ğŸŒ“ AM ç­ï¼ˆä¸Šåˆï¼‰KPI")
        render_kpis([
            KPI("ç¸½äººæ•¸", f"{kpi['total_people']:,}"),
            KPI("é”æ¨™äººæ•¸ï¼ˆæ•´é«”æ•ˆç‡ï¼‰", f"{kpi['total_met']:,}"),
            KPI("é”æ¨™ç‡ï¼ˆæ•´é«”æ•ˆç‡ï¼‰", f"{kpi['total_rate']:.1%}"),
            KPI("é”æ¨™é–€æª»", f"æ•ˆç‡ â‰¥ {TARGET_EFF}"),
        ])
        card_close()

        card_open(f"AM ç­ï¼ˆä¸Šåˆï¼‰æ•ˆç‡æ’è¡Œï¼ˆTop {int(meta['top_n'])}ï¼‰")
        am_rank = summary[[user_col, "å°æ‡‰å§“å", "ä¸Šåˆç­†æ•¸", "ä¸Šåˆå·¥æ™‚_åˆ†é˜", "ä¸Šåˆæ•ˆç‡_ä»¶æ¯å°æ™‚"]].copy()
        am_rank = am_rank.rename(columns={"ä¸Šåˆæ•ˆç‡_ä»¶æ¯å°æ™‚": "æ•ˆç‡", "ä¸Šåˆç­†æ•¸": "ç­†æ•¸", "ä¸Šåˆå·¥æ™‚_åˆ†é˜": "å·¥æ™‚"})
        am_rank["å§“å"] = am_rank["å°æ‡‰å§“å"].where(am_rank["å°æ‡‰å§“å"].astype(str).str.len() > 0, am_rank[user_col].astype(str))
        bar_topN(
            am_rank[["å§“å", "æ•ˆç‡", "ç­†æ•¸", "å·¥æ™‚"]],
            x_col="å§“å", y_col="æ•ˆç‡",
            hover_cols=["ç­†æ•¸", "å·¥æ™‚"],
            top_n=int(meta["top_n"]),
            target=float(TARGET_EFF),
        )
        card_close()

    with col_r:
        card_open("ğŸŒ™ PM ç­ï¼ˆä¸‹åˆï¼‰KPI")
        render_kpis([
            KPI("ç¸½äººæ•¸", f"{kpi['pm_total']:,}"),
            KPI("é”æ¨™äººæ•¸ï¼ˆä¸‹åˆæ•ˆç‡ï¼‰", f"{kpi['pm_met']:,}"),
            KPI("é”æ¨™ç‡ï¼ˆä¸‹åˆæ•ˆç‡ï¼‰", f"{kpi['pm_rate']:.1%}"),
            KPI("é”æ¨™é–€æª»", f"æ•ˆç‡ â‰¥ {TARGET_EFF}"),
        ])
        card_close()

        card_open(f"PM ç­ï¼ˆä¸‹åˆï¼‰æ•ˆç‡æ’è¡Œï¼ˆTop {int(meta['top_n'])}ï¼‰")
        pm_rank = summary[[user_col, "å°æ‡‰å§“å", "ä¸‹åˆç­†æ•¸", "ä¸‹åˆå·¥æ™‚_åˆ†é˜_æ‰£ä¼‘", "ä¸‹åˆæ•ˆç‡_ä»¶æ¯å°æ™‚"]].copy()
        pm_rank = pm_rank.rename(columns={"ä¸‹åˆæ•ˆç‡_ä»¶æ¯å°æ™‚": "æ•ˆç‡", "ä¸‹åˆç­†æ•¸": "ç­†æ•¸", "ä¸‹åˆå·¥æ™‚_åˆ†é˜_æ‰£ä¼‘": "å·¥æ™‚"})
        pm_rank["å§“å"] = pm_rank["å°æ‡‰å§“å"].where(pm_rank["å°æ‡‰å§“å"].astype(str).str.len() > 0, pm_rank[user_col].astype(str))
        bar_topN(
            pm_rank[["å§“å", "æ•ˆç‡", "ç­†æ•¸", "å·¥æ™‚"]],
            x_col="å§“å", y_col="æ•ˆç‡",
            hover_cols=["ç­†æ•¸", "å·¥æ™‚"],
            top_n=int(meta["top_n"]),
            target=float(TARGET_EFF),
        )
        card_close()

    # âœ… åŒ¯å‡ºï¼šä¸€è¡Œ=æŒ‰éˆ•ï¼›æŒ‰ä¸‹å» KPI ä»ä¿ç•™
    card_open("â¬‡ï¸ åŒ¯å‡º KPI å ±è¡¨ï¼ˆExcelï¼‰")
    default_name = f"{meta['source_filename'].rsplit('.', 1)[0]}_ä¸Šæ¶ç¸¾æ•ˆ.xlsx"
    st.download_button(
        label="â¬‡ï¸ åŒ¯å‡º Excelï¼ˆå½™ç¸½/æ˜ç´°/æ™‚æ®µ/è¦å‰‡ï¼‰",
        data=xlsx_bytes,
        file_name=default_name,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )
    card_close()

    # ç•™å­˜ï¼ˆSupabaseï¼‰
    st.divider()
    st.subheader("ğŸ§¾ ç¨½æ ¸ç•™å­˜ç‹€æ…‹ï¼ˆSupabaseï¼‰")
    try:
        row = try_audit_persist()
        if row:
            st.success(f"âœ… å·²ç•™å­˜æœ¬æ¬¡åˆ†æï¼ˆIDï¼š{row.get('id','')}ï¼‰")
        else:
            st.info("æœ¬æ¬¡çµæœå·²ç•™å­˜ï¼ˆæœªé‡è¤‡å¯«å…¥ï¼‰ã€‚")
    except Exception as e:
        st.error("âŒ ç•™å­˜å¤±æ•—ï¼ˆä¸å½±éŸ¿åŒ¯å‡ºèˆ‡ç•«é¢ï¼‰")
        st.code(repr(e))


if __name__ == "__main__":
    main()
