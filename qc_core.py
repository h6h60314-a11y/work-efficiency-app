
#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
驗收達標效率（含空窗＋姓名對照＋休息規則＋空窗明細＋條件著色＋AM/PM 分段）
v18 排除規則版：
- 午休 12:30–13:30 不計空窗（全時段空窗已扣除重疊）
- 新增『午後空窗…』三欄僅用於下午達標，且只在 13:30 後兩筆任務之間才計空窗
  → 「13:30 → 第一筆下午任務」不記入下午空窗
- 下午區塊的空窗筆數/分鐘/明細一律改用『午後空窗…』三欄
- ✅ 可輸入「多名人員＋各自時間區間」排除規則：
  - 這些人＋時間內的紀錄不參與任何統計（筆數、效率、AM/PM…）
  - 這些時間區間也從空窗計算中扣除（不算空窗）
  - 全日、AM/PM 的「總分鐘」都會扣除這些排除時間
"""

from __future__ import annotations
import os
import numpy as np
import pandas as pd
import tempfile
import io
from datetime import datetime, time
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter

# ===== 可調參數 =====
THRESHOLD_MIN = 10  # 空窗門檻（分鐘）
USER_COLS = ["記錄輸入人","建立人員","建立者","輸入人","建立者姓名","操作人員","建立人"]
TIME_COLS = ["修訂日期","更新日期","異動日期","修改日期","最後更新時間","時間戳記","Timestamp"]
DEST_COL = "到"; DEST_VALUE_QC = "QC"

# AM/PM 切段
AM_START = time(9, 0)
AM_END   = time(12, 30)
PM_START = time(13, 30)

# 午休不計空窗時間窗
LUNCH_START = time(12, 30)
LUNCH_END   = time(13, 30)

# === 記錄輸入人 → 姓名 對照（可自行擴充） ===
ID_TO_NAME = {
    "09440": "張予軒","10137": "徐嘉蔆","10818": "葉青芳","11797": "賴泉和",
    "20201109001": "吳振凱","10003": "李茂銓","10471": "余興炫","10275": "羅仲宇",
    "9440": "張予軒",
}
def map_name_from_id(x: str) -> str:
    if x is None: return ""
    s = str(x).strip()
    if not s: return ""
    if s in ID_TO_NAME: return ID_TO_NAME[s]
    return ID_TO_NAME.get(s.lstrip("0"), "")

# ---------- 小工具 ----------
def pick_col(cols, candidates):
    cols_norm = [str(c).strip() for c in cols]
    for cand in candidates:
        if cand in cols_norm: return cand
    for cand in candidates:
        for c in cols_norm:
            if cand in c: return c
    return None

def to_dt(series: pd.Series) -> pd.Series:
    if pd.api.types.is_datetime64_any_dtype(series):
        return pd.to_datetime(series, errors="coerce")
    patterns = [
        "%Y-%m-%d %H:%M:%S","%Y/%m/%d %H:%M:%S",
        "%Y-%m-%d %H:%M","%Y/%m/%d %H:%M",
        "%m/%d/%Y %H:%M","%m/%d/%Y %H:%M:%S",
    ]
    def parse_one(x):
        x = str(x).strip()
        for p in patterns:
            try: return datetime.strptime(x, p)
            except Exception: pass
        return pd.NaT
    return series.apply(parse_one)

def read_any(path: str) -> dict:
    ext = os.path.splitext(path)[1].lower()
    if ext in [".xlsx",".xlsm",".xltx",".xltm"]:
        return pd.read_excel(path, sheet_name=None, engine="openpyxl")
    if ext == ".xls":   # 老 .xls 需 xlrd，可能會有 OLE2 警告，不影響輸出 .xlsx
        return pd.read_excel(path, sheet_name=None)
    if ext in [".csv",".txt"]:
        return {"CSV": pd.read_csv(path, encoding="utf-8", low_memory=False)}
    try:
        return pd.read_excel(path, sheet_name=None, engine="openpyxl")
    except Exception:
        return {"CSV": pd.read_csv(path, encoding="utf-8", low_memory=False)}

# ---------- 計算「排除時間區間」的分鐘數（用在總分鐘） ----------
def calc_exclude_minutes_for_range(date_obj, user_id, first_ts, last_ts, skip_rules):
    """
    計算某人、某日，在 first_ts~last_ts 期間內，
    所有『排除時間區間』的總分鐘數（做聯集合併，避免重複計算）。
    """
    if pd.isna(first_ts) or pd.isna(last_ts) or not skip_rules:
        return 0

    from datetime import datetime as _dt

    segs = []
    user_id_str = str(user_id).strip()

    for rule in skip_rules:
        rule_user = str(rule["user"]).strip()
        # rule_user 有填才比對，空字串代表「全部人」
        if rule_user and rule_user != user_id_str:
            continue

        s_dt = _dt.combine(date_obj, rule["t_start"])
        e_dt = _dt.combine(date_obj, rule["t_end"])

        left  = max(first_ts, s_dt)
        right = min(last_ts, e_dt)
        if right > left:
            segs.append((left, right))

    if not segs:
        return 0

    # 聯集合併
    segs.sort(key=lambda x: x[0])
    merged = [list(segs[0])]
    for s, e in segs[1:]:
        ls, le = merged[-1]
        if s <= le:
            if e > le:
                merged[-1][1] = e
        else:
            merged.append([s, e])

    total = 0.0
    for s, e in merged:
        total += (e - s).total_seconds() / 60.0
    return total

# ---------- 空窗計算（含午休扣除＋排除區間＋『午後空窗』三欄） ----------
def annotate_idle(qc_df: pd.DataFrame, user_col: str, time_col: str, skip_rules=None) -> pd.DataFrame:
    """
    逐人依時間排序：
    - 全時段空窗：
        兩筆間隔扣除：
          1. 午休 12:30–13:30
          2. 所有「排除時間區間」（符合人員 or 全員）
        的重疊後，若 > THRESHOLD_MIN 記為空窗。
    - 『午後空窗』：
        僅在同日且前一筆時間 >= 13:30、下一筆時間 > 前一筆 的間隔，
        同樣扣掉排除時間後，若 > THRESHOLD_MIN 才記；
        → 不會把「13:30 → 當天第一筆下午任務」視為空窗。
    """
    merged = qc_df.copy()
    for col in ["空窗分鐘","空窗旗標","空窗區間","午後空窗分鐘","午後空窗旗標","午後空窗區間"]:
        if col not in merged.columns: merged[col] = pd.NA

    tmp = merged[[user_col, time_col]].copy()
    tmp["_user"] = tmp[user_col].astype(str).str.strip()
    tmp["_dt"] = to_dt(tmp[time_col])
    tmp = tmp.loc[tmp["_dt"].notna()].copy()
    tmp.sort_values(by=["_user","_dt"], inplace=True)
    tmp["_prev_dt"] = tmp.groupby("_user")["_dt"].shift(1)

    idle_minutes, idle_flag, idle_text = [], [], []
    pm_minutes, pm_flag, pm_text = [], [], []
    from datetime import datetime as _dt

    if skip_rules is None:
        skip_rules = []

    for _, r in tmp.iterrows():
        prev_dt = r["_prev_dt"]; cur_dt = r["_dt"]
        user_id = r["_user"]
        if pd.isna(prev_dt):
            idle_minutes.append(np.nan); idle_flag.append(0); idle_text.append("")
            pm_minutes.append(np.nan); pm_flag.append(0); pm_text.append("")
            continue

        # 原始間隔
        gap_min = (cur_dt - prev_dt).total_seconds() / 60.0

        # 準備需要扣除的時間區段（午休 + 自訂排除區間），統一做「聯集」計算
        segments = []
        if prev_dt.date() == cur_dt.date():
            # 午休
            lunch_start_dt = _dt.combine(cur_dt.date(), LUNCH_START)
            lunch_end_dt   = _dt.combine(cur_dt.date(), LUNCH_END)
            left  = max(prev_dt, lunch_start_dt)
            right = min(cur_dt, lunch_end_dt)
            if right > left:
                segments.append((left, right))

            # 自訂排除區間（依人員＋時間）
            for rule in skip_rules:
                rule_user = str(rule["user"]).strip()
                # 人員符合（空字串代表全員）、且同一天
                if rule_user and rule_user != user_id:
                    continue
                exc_start_dt = _dt.combine(cur_dt.date(), rule["t_start"])
                exc_end_dt   = _dt.combine(cur_dt.date(), rule["t_end"])
                left  = max(prev_dt, exc_start_dt)
                right = min(cur_dt, exc_end_dt)
                if right > left:
                    segments.append((left, right))

        # 計算所有 segments 的聯集總長度（分鐘）
        overlap_min = 0.0
        if segments:
            segments.sort(key=lambda x: x[0])
            merged_seg = [list(segments[0])]
            for s, e in segments[1:]:
                last_s, last_e = merged_seg[-1]
                if s <= last_e:
                    if e > last_e:
                        merged_seg[-1][1] = e
                else:
                    merged_seg.append([s, e])
            for s, e in merged_seg:
                overlap_min += (e - s).total_seconds() / 60.0

        eff_gap = gap_min - overlap_min  # 已扣午休 + 排除區間 的有效空窗

        # === 全時段空窗（供全日統計與上午用） ===
        if eff_gap > THRESHOLD_MIN:
            idle_minutes.append(int(eff_gap)); idle_flag.append(1)
            idle_text.append(f"{prev_dt.strftime('%H:%M')} ~ {cur_dt.strftime('%H:%M')}")
        else:
            idle_minutes.append(np.nan); idle_flag.append(0); idle_text.append("")

        # === 僅 13:30 以後兩筆之間（下午空窗）：prev >= 13:30 且同日 ===
        pm_min = 0.0
        pm_segs = []
        if prev_dt.date() == cur_dt.date() and prev_dt.time() >= LUNCH_END:
            # 一樣要扣掉「午休 + 排除區間」的時間
            pm_gap_min = (cur_dt - prev_dt).total_seconds() / 60.0 - overlap_min
            if pm_gap_min > 0:
                pm_min = pm_gap_min
                pm_segs.append(f"{prev_dt.strftime('%H:%M')} ~ {cur_dt.strftime('%H:%M')}")

        if pm_min > THRESHOLD_MIN:
            pm_minutes.append(int(pm_min)); pm_flag.append(1); pm_text.append("、".join(pm_segs))
        else:
            pm_minutes.append(np.nan); pm_flag.append(0); pm_text.append("")

    idx = tmp.index
    merged.loc[idx, "空窗分鐘"]   = idle_minutes
    merged.loc[idx, "空窗旗標"]   = idle_flag
    merged.loc[idx, "空窗區間"]   = idle_text
    merged.loc[idx, "午後空窗分鐘"] = pm_minutes
    merged.loc[idx, "午後空窗旗標"] = pm_flag
    merged.loc[idx, "午後空窗區間"] = pm_text
    return merged

# ---------- 休息規則 ----------
from datetime import time as _Time
def _t(h,m): return _Time(hour=h, minute=m)
def _ge(t:_Time, x:_Time): return (t is not None) and (t >= x)
def _le(t:_Time, x:_Time): return (t is not None) and (t <= x)

def calc_rest_minutes_for_day(first_ts: pd.Timestamp, last_ts: pd.Timestamp) -> int:
    if pd.isna(first_ts) or pd.isna(last_ts): return 0
    f, l = first_ts.time(), last_ts.time()
    if _ge(f,_t(13,29)) and _le(l,_t(15,29)): return 0    # E
    if _ge(f,_t(17,0))  and _le(l,_t(18,29)): return 0    # F
    if _ge(f,_t(18,0))  and _le(l,_t(20,29)): return 0    # I
    if _ge(f,_t(17,30)) and _le(l,_t(20,40)): return 30   # K
    if _ge(f,_t(10,29)) and _le(l,_t(15,29)): return 60   # A
    if _ge(f,_t(9,0))   and _le(l,_t(15,45)): return 75   # N
    if _ge(f,_t(9,0))   and _le(l,_t(16,0)):  return 90   # B
    if _ge(f,_t(10,30)) and _le(l,_t(17,0)):  return 75   # L
    if _ge(f,_t(9,0))   and _le(l,_t(18,15)): return 90   # D
    if _ge(f,_t(9,0))   and _le(l,_t(17,59)): return 90   # C
    if _ge(f,_t(10,30)) and _le(l,_t(20,40)): return 105  # J
    if _ge(f,_t(9,0))   and _le(l,_t(20,29)): return 120  # G
    if _ge(f,_t(9,0))   and _le(l,_t(21,0)):  return 135  # H
    if _ge(f,_t(9,0))   and _le(l,_t(21,10)): return 135  # M
    return 0

def _within_am(dt: pd.Timestamp) -> bool:
    t = dt.time()
    return (t >= AM_START) and (t <= AM_END)
def _within_pm(dt: pd.Timestamp) -> bool:
    t = dt.time()
    return (t >= PM_START)

def calc_rest_minutes_for_pm(first_ts: pd.Timestamp, last_ts: pd.Timestamp) -> int:
    if pd.isna(first_ts) or pd.isna(last_ts): return 0
    f, l = first_ts.time(), last_ts.time()
    # 優先序 E → F → I → J → K
    if _ge(f,_t(13,29)) and _le(l,_t(15,29)): return 0
    if _ge(f,_t(13,29)) and _le(l,_t(18,0)):  return 15
    if _ge(f,_t(13,29)) and _le(l,_t(19,59)): return 15
    if _ge(f,_t(13,29)) and _le(l,_t(20,39)): return 45
    if _ge(f,_t(13,29)) and _le(l,_t(20,40)): return 45
    return 0

# ---------- 全日統計 ----------
def build_efficiency_table_full(qc_with_idle: pd.DataFrame, user_col: str, time_col: str, skip_rules=None) -> pd.DataFrame:
    if skip_rules is None:
        skip_rules = []

    df = qc_with_idle.copy()
    df["_user"] = df[user_col].astype(str).str.strip()
    df["_name"] = df["_user"].apply(map_name_from_id)
    df["_dt"]   = to_dt(df[time_col])
    df = df.loc[df["_dt"].notna()].copy()
    df["_date"] = df["_dt"].dt.date

    agg_size = df.groupby(["_date","_user","_name"]).size().reset_index(name="筆數")
    times = df.groupby(["_date","_user","_name"])["_dt"].agg(第一筆修訂日期="min", 最後一筆修訂日期="max").reset_index()
    out = agg_size.merge(times, on=["_date","_user","_name"], how="left")

    # 休息分鐘
    out["休息分鐘"] = out.apply(
        lambda r: calc_rest_minutes_for_day(r["第一筆修訂日期"], r["最後一筆修訂日期"]),
        axis=1
    )

    # 原始總分鐘（未扣休息、未扣排除）
    total_min_raw = (out["最後一筆修訂日期"] - out["第一筆修訂日期"]).dt.total_seconds().div(60)

    # 要扣除的「排除時間區間」分鐘
    exclude_minutes = out.apply(
        lambda r: calc_exclude_minutes_for_range(
            r["_date"],          # 還沒 rename 前
            r["_user"],
            r["第一筆修訂日期"],
            r["最後一筆修訂日期"],
            skip_rules
        ),
        axis=1
    )

    out["總分鐘"] = total_min_raw - out["休息分鐘"] - exclude_minutes
    out.loc[out["總分鐘"] <= 0, "總分鐘"] = np.nan
    out["總工時"] = out["總分鐘"] / 60
    out["效率"]   = out["筆數"] / out["總工時"]

    gap_flag = df["空窗旗標"] if "空窗旗標" in df.columns else 0
    idle_count = (df.assign(旗=gap_flag)
                    .groupby(["_date","_user"])["旗"].sum().reset_index().rename(columns={"旗":"空窗筆數"}))
    gap_minutes = (df.groupby(["_date","_user"])["空窗分鐘"].sum()
                    .reset_index().rename(columns={"空窗分鐘":"空窗總分鐘"}))
    gap_text = (df.groupby(["_date","_user"])["空窗區間"]
                  .apply(lambda s: "、".join([x for x in s if isinstance(x,str) and x.strip()]))
                  .reset_index().rename(columns={"空窗區間":"空窗明細"}))

    out = (out.merge(idle_count, on=["_date","_user"], how="left")
              .merge(gap_minutes, on=["_date","_user"], how="left")
              .merge(gap_text, on=["_date","_user"], how="left"))

    out.rename(columns={"_date":"日期","_user":"記錄輸入人","_name":"姓名"}, inplace=True)
    out["空窗筆數"]   = out["空窗筆數"].fillna(0).astype(int)
    out["空窗總分鐘"] = out["空窗總分鐘"].fillna(0).astype(int)
    out["空窗明細"]   = out["空窗明細"].fillna("")

    out["總分鐘"] = out["總分鐘"].round(2)
    out["總工時"] = out["總工時"].round(2)
    out["效率"]   = out["效率"].round(2)

    col_order = [
        "日期","記錄輸入人","姓名","筆數",
        "第一筆修訂日期","最後一筆修訂日期",
        "休息分鐘","總分鐘","總工時","效率",
        "空窗筆數","空窗總分鐘","空窗明細"
    ]
    return out[col_order].sort_values(by=["日期","記錄輸入人","第一筆修訂日期"])

# ---------- AM/PM 分段（下午用『午後空窗…』） ----------
def build_efficiency_table_ampm(qc_with_idle: pd.DataFrame, user_col: str, time_col: str, skip_rules=None) -> pd.DataFrame:
    if skip_rules is None:
        skip_rules = []

    df = qc_with_idle.copy()
    df["_user"] = df[user_col].astype(str).str.strip()
    df["_name"] = df["_user"].apply(map_name_from_id)
    df["_dt"]   = to_dt(df[time_col])
    df = df.loc[df["_dt"].notna()].copy()
    df["_date"] = df["_dt"].dt.date
    df.sort_values(by=["_user","_dt"], inplace=True)

    out_rows = []
    for (d, u, n), g in df.groupby(["_date","_user","_name"]):
        g_am = g.loc[g["_dt"].apply(_within_am)].copy()
        g_pm = g.loc[g["_dt"].apply(_within_pm)].copy()

        def make_row(sub: pd.DataFrame, label: str):
            if sub.empty: return
            first_ts = sub["_dt"].min()
            last_ts  = sub["_dt"].max()
            rest_min = 15 if label=="上午" else calc_rest_minutes_for_pm(first_ts, last_ts)
            total_min_raw = (last_ts - first_ts).total_seconds()/60

            # 同一個日期 d、同一個人 u，在這個 AM / PM 時段內要扣的排除分鐘
            exclude_min = calc_exclude_minutes_for_range(
                d,      # 這個 group 的日期
                u,      # 記錄輸入人
                first_ts,
                last_ts,
                skip_rules
            )

            total_min = total_min_raw - rest_min - exclude_min
            count = len(sub)
            if total_min <= 0 or pd.isna(total_min):
                total_hr = np.nan; eff = np.nan
            else:
                total_hr = total_min/60; eff = count/total_hr if total_hr else np.nan

            if label == "下午":
                idle_cnt = int(sub["午後空窗旗標"].fillna(0).sum()) if "午後空窗旗標" in sub.columns else 0
                idle_min = int(sub["午後空窗分鐘"].fillna(0).sum()) if "午後空窗分鐘" in sub.columns else 0
                idle_text_items = [x for x in sub.get("午後空窗區間", pd.Series(dtype=object)).tolist()
                                   if isinstance(x,str) and x.strip()]
            else:
                idle_cnt = int(sub["空窗旗標"].fillna(0).sum()) if "空窗旗標" in sub.columns else 0
                idle_min = int(sub["空窗分鐘"].fillna(0).sum()) if "空窗分鐘" in sub.columns else 0
                idle_text_items = [x for x in sub.get("空窗區間", pd.Series(dtype=object)).tolist()
                                   if isinstance(x,str) and x.strip()]
            out_rows.append({
                "日期": d, "時段": label, "記錄輸入人": u, "姓名": n, "筆數": count,
                "第一筆修訂日期": first_ts, "最後一筆修訂日期": last_ts,
                "休息分鐘": rest_min, "總分鐘": total_min, "總工時": total_hr, "效率": eff,
                "空窗筆數": idle_cnt, "空窗總分鐘": idle_min, "空窗明細": "、".join(idle_text_items)
            })

        make_row(g_am, "上午"); make_row(g_pm, "下午")

    if not out_rows:
        return pd.DataFrame(columns=["日期","時段","記錄輸入人","姓名","筆數","第一筆修訂日期","最後一筆修訂日期",
                                     "休息分鐘","總分鐘","總工時","效率","空窗筆數","空窗總分鐘","空窗明細"])
    out = pd.DataFrame(out_rows)
    out["總分鐘"] = out["總分鐘"].round(2)
    out["總工時"] = out["總工時"].round(2)
    out["效率"]   = out["效率"].round(2)
    col_order = ["日期","時段","記錄輸入人","姓名","筆數",
                 "第一筆修訂日期","最後一筆修訂日期",
                 "休息分鐘","總分鐘","總工時","效率",
                 "空窗筆數","空窗總分鐘","空窗明細"]
    return out[col_order].sort_values(by=["日期","記錄輸入人","時段","第一筆修訂日期"])

# ---------- 視覺化：每日期一大標題，上午/下午兩區塊 ----------
def write_grouped_ampm_sheet(wb, ampm_df: pd.DataFrame, sheet_name="AMPM_日期分組"):
    COLS = ["記錄輸入人","姓名","筆數","第一筆修訂日期","最後一筆修訂日期",
            "休息分鐘","總分鐘","總工時","效率","空窗筆數","空窗總分鐘","空窗明細"]
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    title_font = Font(size=14, bold=True)
    header_font = Font(size=11, bold=True)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(top=thin, bottom=thin, left=thin, right=thin)
    green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red   = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    gray  = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    widths = [12,12,7,19,19,9,9,9,8,9,10,60]
    for i,w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    r = 1
    for d, gdate in ampm_df.groupby("日期", sort=True):
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(COLS))
        ws.cell(r,1).value = str(d); ws.cell(r,1).font = title_font
        ws.cell(r,1).alignment = left; ws.cell(r,1).fill = gray
        r += 1

        for label, title in [("上午","上午達標"), ("下午","下午達標")]:
            sub = gdate[gdate["時段"]==label].copy()

            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(COLS))
            ws.cell(r,1).value = title; ws.cell(r,1).font = header_font
            ws.cell(r,1).alignment = left
            r += 1

            # 表頭
            for c_idx, col in enumerate(COLS, start=1):
                ws.cell(r,c_idx).value = col
                ws.cell(r,c_idx).font = header_font
                ws.cell(r,c_idx).alignment = center
                ws.cell(r,c_idx).border = Border(top=thin,bottom=thin,left=thin,right=thin)
            r += 1

            # 明細
            if not sub.empty:
                sub = sub[COLS]
                for _, row in sub.iterrows():
                    for c_idx, col in enumerate(COLS, start=1):
                        cell = ws.cell(r,c_idx, row[col])
                        cell.alignment = left if col == "空窗明細" else center
                        cell.border = Border(top=thin,bottom=thin,left=thin,right=thin)
                    eff = row["效率"]
                    fill = green if (pd.notna(eff) and eff >= 20) else red
                    for c_idx in range(1, len(COLS)+1):
                        ws.cell(r,c_idx).fill = fill
                    ws.cell(r, COLS.index("總分鐘")+1).number_format = "0.00"
                    ws.cell(r, COLS.index("總工時")+1).number_format = "0.00"
                    ws.cell(r, COLS.index("效率")+1).number_format   = "0.00"
                    r += 1
            else:
                ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(COLS))
                ws.cell(r,1).value = "(無資料)"; ws.cell(r,1).alignment = left
                r += 1

            r += 1  # 區塊間空一行
        r += 1      # 每日間空一行

# ===================== Streamlit/Cloud 可呼叫入口 =====================
def run_qc_efficiency(file_bytes: bytes, original_name: str, skip_rules: list[dict] | None = None) -> dict:
    """
    Streamlit / API 入口：上傳檔(bytes) → 回傳統計表 + 已格式化的 Excel(bytes)

    Parameters
    ----------
    file_bytes : bytes
        上傳檔案內容（Excel/CSV）
    original_name : str
        原始檔名（用來判斷副檔名）
    skip_rules : list[dict] | None
        排除規則（可多筆）：
        [
          {"user": "20201109001" 或 ""(空字串=全員), "t_start": datetime.time, "t_end": datetime.time},
          ...
        ]

    Returns
    -------
    dict:
      {
        "full_df": DataFrame,   # 記錄輸入人統計（全日）
        "ampm_df": DataFrame,   # 記錄輸入人統計（AM/PM）
        "idle_df": DataFrame,   # 空窗明細
        "xlsx_bytes": bytes,    # 含條件著色+AMPM日期分組的輸出 Excel
        "total_idle": int,      # 全體空窗筆數
      }
    """
    if skip_rules is None:
        skip_rules = []

    # 基本清理：確保 user 是字串、時間是 time
    cleaned = []
    for r in skip_rules:
        if not isinstance(r, dict):
            continue
        user = str(r.get("user", "")).strip()
        t_start = r.get("t_start", None)
        t_end = r.get("t_end", None)
        if t_start is None or t_end is None:
            continue
        # st.time_input 會回傳 datetime.time；若是字串也容錯
        if isinstance(t_start, str):
            t_start = datetime.strptime(t_start, "%H:%M").time()
        if isinstance(t_end, str):
            t_end = datetime.strptime(t_end, "%H:%M").time()
        if t_end < t_start:
            continue
        cleaned.append({"user": user, "t_start": t_start, "t_end": t_end})
    skip_rules = cleaned

    suffix = os.path.splitext(original_name)[1].lower()
    if suffix not in [".xlsx", ".xlsm", ".xls", ".csv", ".txt", ".xltx", ".xltm"]:
        suffix = ".xlsx"

    processed = {}
    idle_details_all = []

    with tempfile.TemporaryDirectory() as td:
        in_path = os.path.join(td, f"upload{suffix}")
        with open(in_path, "wb") as f:
            f.write(file_bytes)

        sheets = read_any(in_path)

        # 2) 每張表處理：找 QC，算空窗，補姓名（保留你原本邏輯）
        for name, df in sheets.items():
            if df is None or df.empty:
                processed[name] = df
                continue

            df = df.copy()
            dest_col = pick_col(df.columns, [DEST_COL])
            if dest_col and DEST_VALUE_QC in df[dest_col].astype(str).unique().tolist():
                qc = df.loc[df[dest_col].astype(str) == DEST_VALUE_QC].copy()
            else:
                qc = df.copy()

            ucol = pick_col(qc.columns, USER_COLS)
            tcol = pick_col(qc.columns, TIME_COLS)

            # ====== 先排除「多筆人員＋時間區間」的紀錄（不參與任何統計） ======
            if ucol and tcol and skip_rules:
                dt_series = to_dt(qc[tcol])
                t_series = dt_series.dt.time

                mask_all = pd.Series(False, index=qc.index)
                for rule in skip_rules:
                    t_start = rule["t_start"]
                    t_end = rule["t_end"]
                    user_rule = str(rule["user"]).strip()

                    def _time_in_range(t, ts=t_start, te=t_end):
                        return isinstance(t, time) and (t >= ts) and (t <= te)

                    mask_time = t_series.apply(_time_in_range)
                    if user_rule:
                        mask_user = qc[ucol].astype(str).str.strip() == user_rule
                    else:
                        mask_user = pd.Series(True, index=qc.index)

                    mask_all = mask_all | (mask_time & mask_user)

                exclude_idx = qc.index[mask_all]
                if len(exclude_idx) > 0:
                    qc = qc.drop(exclude_idx)
                    df = df.drop(exclude_idx, errors="ignore")

            # ====== 欄位不齊就補空窗欄/姓名後直接輸出 ======
            if not ucol or not tcol:
                for col in ["空窗分鐘", "空窗旗標", "空窗區間", "午後空窗分鐘", "午後空窗旗標", "午後空窗區間"]:
                    if col not in df.columns:
                        df[col] = pd.NA
                user_guess = pick_col(df.columns, USER_COLS)
                if user_guess and "姓名" not in df.columns:
                    df["姓名"] = df[user_guess].astype(str).apply(map_name_from_id)
                processed[name] = df
                continue

            # 空窗計算會再扣掉：午休 + 「排除區間」時間（你的 annotate_idle 已支援）
            qc_with_idle = annotate_idle(qc, ucol, tcol, skip_rules=skip_rules)

            df_out = df.copy()
            df_out.loc[qc_with_idle.index, ["空窗分鐘","空窗旗標","空窗區間",
                                            "午後空窗分鐘","午後空窗旗標","午後空窗區間"]] = \
               qc_with_idle[["空窗分鐘","空窗旗標","空窗區間",
                             "午後空窗分鐘","午後空窗旗標","午後空窗區間"]].values

            if "姓名" not in df_out.columns:
                df_out["姓名"] = ""
            try:
                df_out.loc[:, "姓名"] = df_out[ucol].astype(str).apply(map_name_from_id)
            except Exception:
                pass
            processed[name] = df_out

            # 空窗明細分頁資料（上午：空窗旗標；下午：午後空窗旗標）
            if not qc_with_idle.empty:
                tmp = qc_with_idle.copy()
                tmp["_user"] = tmp[ucol].astype(str).str.strip()
                tmp["_name"] = tmp["_user"].apply(map_name_from_id)
                tmp["_dt"]   = to_dt(tmp[tcol])
                tmp = tmp.loc[tmp["_dt"].notna()].copy()
                tmp.sort_values(by=["_user","_dt"], inplace=True)
                tmp["日期"] = tmp["_dt"].dt.date
                tmp["起"] = tmp["_dt"].shift(1).dt.strftime("%H:%M")
                tmp["迄"] = tmp["_dt"].dt.strftime("%H:%M")
                tmp["來源分頁"] = name
                tmp["記錄輸入人"] = tmp["_user"]; tmp["姓名"] = tmp["_name"]

                tmp_am = tmp.loc[tmp["空窗旗標"]==1, ["來源分頁","日期","記錄輸入人","姓名","起","迄","空窗分鐘","空窗區間"]]
                tmp_pm = tmp.loc[tmp["午後空窗旗標"]==1, ["來源分頁","日期","記錄輸入人","姓名","起","迄"]].assign(
                    空窗分鐘=tmp.loc[tmp["午後空窗旗標"]==1,"午後空窗分鐘"].values,
                    空窗區間=tmp.loc[tmp["午後空窗旗標"]==1,"午後空窗區間"].values
                )
                tmp2 = pd.concat([tmp_am, tmp_pm], ignore_index=True)
                if not tmp2.empty:
                    idle_details_all.append(tmp2)

        # 3) 彙整全日/AMPM 表
        full_df = pd.DataFrame()
        ampm_df = pd.DataFrame()
        if processed:
            big = pd.concat(processed.values(), ignore_index=True)
            ucol_all = pick_col(big.columns, USER_COLS)
            tcol_all = pick_col(big.columns, TIME_COLS)
            if ucol_all and tcol_all:
                full_df = build_efficiency_table_full(big, ucol_all, tcol_all, skip_rules=skip_rules)
                ampm_df = build_efficiency_table_ampm(big, ucol_all, tcol_all, skip_rules=skip_rules)

        # 空窗明細彙整 + 排序
        if idle_details_all:
            idle_details = pd.concat(idle_details_all, ignore_index=True)
            final_cols = ["來源分頁","日期","記錄輸入人","姓名","起","迄","空窗分鐘","空窗區間"]
            for c in final_cols:
                if c not in idle_details.columns:
                    idle_details[c] = "" if c in ["來源分頁","記錄輸入人","姓名","起","迄","空窗區間"] else 0
            idle_details = idle_details[final_cols].copy()
            idle_details.sort_values(by=["日期","記錄輸入人","起","迄"], inplace=True, ignore_index=True)
        else:
            idle_details = pd.DataFrame(columns=["來源分頁","日期","記錄輸入人","姓名","起","迄","空窗分鐘","空窗區間"])

        # ===== 一致過濾：只保留「同時有 記錄輸入人 + 姓名」的資料（KPI/圖表/匯出 Excel 全部一致）=====

        def _nonempty_series(s: pd.Series) -> pd.Series:

            return s.fillna("").astype(str).str.strip().ne("")


        def _filter_user_and_name(df: pd.DataFrame) -> pd.DataFrame:

            if df is None or df.empty:

                return df

            if "記錄輸入人" in df.columns and "姓名" in df.columns:

                return df[_nonempty_series(df["記錄輸入人"]) & _nonempty_series(df["姓名"])].copy()

            return df


        full_df = _filter_user_and_name(full_df)

        ampm_df = _filter_user_and_name(ampm_df)

        idle_details = _filter_user_and_name(idle_details)

        total_idle = int(idle_details["空窗分鐘"].notna().sum()) if not idle_details.empty else 0
        total_df = pd.DataFrame({"項目":[f"全體空窗筆數(>{THRESHOLD_MIN}分)"], "數量":[total_idle]})

        # ===== 輸出（保留條件著色 + AMPM_日期分組）=====
        out_path = os.path.join(td, "驗收達標_含空窗_AMPM.xlsx")

        def set_two_decimal_format(ws, col_letter, nrows):
            for r in range(2, nrows+1):
                ws[f"{col_letter}{r}"].number_format = "0.00"

        with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
            # 各來源分頁（含空窗欄）
            for name, df in processed.items():
                safe = (name or "Sheet1")[:31]
                df.to_excel(writer, index=False, sheet_name=safe)

            # 記錄輸入人統計（全日）
            if not full_df.empty:
                full_df.to_excel(writer, index=False, sheet_name="記錄輸入人統計")
                ws = writer.book["記錄輸入人統計"]
                nrows = len(full_df)
                if nrows > 0:
                    data_range = f"A2:M{nrows+1}"
                    eff_col_anchor = "$J2"
                    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    red_fill   = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    ws.conditional_formatting.add(data_range, FormulaRule(
                        formula=[f"=AND({eff_col_anchor}>=20,NOT(ISBLANK({eff_col_anchor})))"],
                        stopIfTrue=False, fill=green_fill))
                    ws.conditional_formatting.add(data_range, FormulaRule(
                        formula=[f"=AND({eff_col_anchor}<20,NOT(ISBLANK({eff_col_anchor})))"],
                        stopIfTrue=False, fill=red_fill))
                    set_two_decimal_format(ws,"H",nrows)
                    set_two_decimal_format(ws,"I",nrows)
                    set_two_decimal_format(ws,"J",nrows)

            # 記錄輸入人統計_AMPM（分段；下午用『午後空窗…』）
            if not ampm_df.empty:
                ampm_df.to_excel(writer, index=False, sheet_name="記錄輸入人統計_AMPM")
                ws2 = writer.book["記錄輸入人統計_AMPM"]
                nrows2 = len(ampm_df)
                if nrows2 > 0:
                    data_range2 = f"A2:N{nrows2+1}"
                    eff_col_anchor2 = "$K2"
                    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    red_fill   = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    ws2.conditional_formatting.add(data_range2, FormulaRule(
                        formula=[f"=AND({eff_col_anchor2}>=20,NOT(ISBLANK({eff_col_anchor2})))"],
                        stopIfTrue=False, fill=green_fill))
                    ws2.conditional_formatting.add(data_range2, FormulaRule(
                        formula=[f"=AND({eff_col_anchor2}<20,NOT(ISBLANK({eff_col_anchor2})))"],
                        stopIfTrue=False, fill=red_fill))
                    set_two_decimal_format(ws2,"I",nrows2)
                    set_two_decimal_format(ws2,"J",nrows2)
                    set_two_decimal_format(ws2,"K",nrows2)

            # 空窗明細 / 總結
            idle_details.to_excel(writer, index=False, sheet_name="空窗明細")
            total_df.to_excel(writer, index=False, sheet_name="空窗統計_總結")

            # 視覺化分頁：AMPM_日期分組
            write_grouped_ampm_sheet(writer.book, ampm_df, sheet_name="AMPM_日期分組")

        # AMPM 分頁文字替換（保留原本功能）
        from openpyxl import load_workbook
        def _rename_ampm_titles(xlsx_path, sheet_name="記錄輸入人統計_AMPM"):
            try:
                wb = load_workbook(xlsx_path)
            except Exception:
                return
            if sheet_name not in wb.sheetnames:
                wb.close(); return
            ws = wb[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    v = cell.value
                    if isinstance(v, str):
                        t = v.strip()
                        if t == "第一階段": cell.value = "上午達標"
                        elif t == "第二階段": cell.value = "下午達標"
            wb.save(xlsx_path); wb.close()

        _rename_ampm_titles(out_path)

        with open(out_path, "rb") as f:
            xlsx_bytes = f.read()

    return {
        "full_df": full_df,
        "ampm_df": ampm_df,
        "idle_df": idle_details,
        "xlsx_bytes": xlsx_bytes,
        "total_idle": total_idle,
    }
