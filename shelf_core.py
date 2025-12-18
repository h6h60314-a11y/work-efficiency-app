#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
各課組產能 - 總上組（上架） core for Streamlit
來源邏輯：由使用者提供的 v8.9 腳本改為可呼叫函式（不使用 tkinter 互動）
- 保留：彙總、明細、明細_時段、報表_區塊、休息規則 分頁
- 保留：效率著色（>=TARGET_EFF 綠；<TARGET_EFF 紅）
- 保留：空窗計算（>=10 分鐘；扣固定排除帶；上午不扣休、下午工時計算扣休）
"""
from __future__ import annotations

import os, re, tempfile, datetime as dt
from typing import Dict, Any, Tuple, List

import pandas as pd

# ====== 參數（可被呼叫端覆寫） ======
TO_EXCLUDE_KEYWORDS = ["CGS", "JCPL", "QC99", "GREAT0001X", "GX010", "PD99"]
TO_EXCLUDE_PATTERN = re.compile("|".join(re.escape(k) for k in TO_EXCLUDE_KEYWORDS), flags=re.IGNORECASE)

INPUT_USER_CANDIDATES = ["記錄輸入人", "記錄輸入者", "建立人", "輸入人"]
REV_DT_CANDIDATES    = ["修訂日期", "修訂時間", "修訂日", "異動時間", "修改時間"]

DEFAULT_TARGET_EFF = 20
DEFAULT_IDLE_MIN_THRESHOLD = 10

AM_START, AM_END = dt.time(7,0,0), dt.time(12,30,0)
PM_START, PM_END = dt.time(13,30,0), dt.time(23,59,59)

NAME_MAP = {
    "20200924001":"黃雅君","20210805001":"郭中合","20220505002":"阮文青明",
    "20221221001":"阮文全","20221222005":"謝忠龍","20230119001":"陶春青",
    "20240926001":"陳莉娜","20241011002":"林雙慧","20250502001":"吳詩敏",
    "20250617001":"阮文譚","20250617003":"喬家寶","20250901009":"張寶萱",
    "G01":"0","20201109003":"吳振凱","09963":"黃謙凱",
    "20240313003":"阮曰忠","20201109001":"梁冠如","10003":"李茂銓",
    "20200922002":"葉欲弘","20250923019":"阮氏紅深","9963":"黃謙凱",
    "11399":"陳哲沅",
}

BREAK_RULES = [
     (dt.time(20,45,0), dt.time(22,30,0),  0, "首≥20:45 且 末≤22:30 → 0 分鐘"),
     (dt.time(18,30,0), dt.time(20,30,0),  0, "首≥18:30 且 末≤20:30 → 0 分鐘"),
     (dt.time(15,30,0), dt.time(18, 0,0),  0, "首≥15:30 且 末≤18:00 → 0 分鐘"),
     (dt.time(13,30,0), dt.time(15,35,0),  0, "首≥13:30 且 末≤15:35 → 0 分鐘"),
     (dt.time(20,45,0), dt.time(23, 0,0),  0, "首≥20:45 且 末≤23:00 → 0 分鐘"),
     (dt.time(20, 0,0), dt.time(22, 0,0), 15, "首≥20:00 且 末≤22:00 → 15 分鐘"),
     (dt.time(18,30,0), dt.time(22, 0,0), 15, "首≥18:30 且 末≤22:00 → 15 分鐘"),
     (dt.time(19, 0,0), dt.time(22,30,0), 15, "首≥19:00 且 末≤22:30 → 15 分鐘"),
     (dt.time(13,30,0), dt.time(18, 0,0), 15, "首≥13:30 且 末≤18:00 → 15 分鐘"),
     (dt.time(16, 0,0), dt.time(20,40,0), 30, "首≥16:00 且 末≤20:40 → 30 分鐘"),
     (dt.time(15,30,0), dt.time(20,30,0), 30, "首≥15:30 且 末≤20:30 → 30 分鐘"),
     (dt.time(17, 0,0), dt.time(22,30,0), 45, "首≥17:00 且 末≤22:30 → 45 分鐘"),
     (dt.time(15,45,0), dt.time(22,30,0), 45, "首≥15:45 且 末≤22:30 → 45 分鐘"),
     (dt.time(13,30,0), dt.time(20,29,0), 45, "首≥13:30 且 末≤20:29 → 45 分鐘"),
     (dt.time(13,30,0), dt.time(23, 0,0), 60, "首≥13:30 且 末≤23:00 → 60 分鐘"),
     (dt.time(11, 0,0), dt.time(17, 0,0), 75, "首≥11:00 且 末≤17:00 → 75 分鐘"),
     (dt.time( 8, 0,0), dt.time(17, 0,0), 90, "首≥08:00 且 末≤17:00 → 90 分鐘"),
     (dt.time(10,50,0), dt.time(23, 0,0),120, "首≥10:50 且 末≤23:00 → 120 分鐘"),
     (dt.time( 8, 0,0), dt.time(23, 0,0),135, "首≥08:00 且 末≤23:00 → 135 分鐘"),
]

EXCLUDE_IDLE_RANGES = [
    (dt.time(10, 0, 0), dt.time(10, 15, 0)),
    (dt.time(12,30, 0), dt.time(13, 30, 0)),
    (dt.time(15,30, 0), dt.time(15, 45, 0)),
    (dt.time(18, 0, 0), dt.time(18, 30, 0)),
    (dt.time(20,30, 0), dt.time(20, 45, 0)),
]

def _strip_cols(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.columns = [str(c).strip() for c in df.columns]
    return df

def find_first_column(df: pd.DataFrame, candidates: List[str]) -> str | None:
    cols = {str(c).strip(): c for c in df.columns}
    for name in candidates:
        if name in cols:
            return name
    norm = {re.sub(r"[（）\(\)\s]", "", k): k for k in cols}
    for name in candidates:
        key = re.sub(r"[（）\(\)\s]", "", name)
        if key in norm:
            return norm[key]
    return None

def read_excel_any_quiet(path: str) -> Dict[str, pd.DataFrame]:
    ext = os.path.splitext(path)[1].lower()
    if ext in (".xlsx", ".xlsm"):
        xl = pd.ExcelFile(path, engine="openpyxl")
        return {sn: pd.read_excel(xl, sheet_name=sn) for sn in xl.sheet_names}
    if ext == ".xls":
        xl = pd.ExcelFile(path, engine="xlrd")
        return {sn: pd.read_excel(xl, sheet_name=sn) for sn in xl.sheet_names}
    if ext == ".xlsb":
        xl = pd.ExcelFile(path, engine="pyxlsb")
        return {sn: pd.read_excel(xl, sheet_name=sn) for sn in xl.sheet_names}
    if ext == ".csv":
        for enc in ("utf-8-sig", "cp950", "big5"):
            try:
                return {"CSV": pd.read_csv(path, encoding=enc)}
            except Exception:
                continue
        raise Exception("CSV 讀取失敗。")
    raise Exception("不支援的副檔名。")

def normalize_to_qc(series: pd.Series) -> pd.Series:
    s = series.astype(str).str.strip().str.upper()
    return s.eq("QC")

def to_not_excluded_mask(series: pd.Series) -> pd.Series:
    s = series.astype(str).str.strip()
    return ~s.str.contains(TO_EXCLUDE_PATTERN, na=False)

def prepare_filtered_df(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame()
    df = _strip_cols(df)
    if "由" not in df.columns or "到" not in df.columns:
        return pd.DataFrame()
    return df[normalize_to_qc(df["由"]) & to_not_excluded_mask(df["到"])].copy()

def autosize_columns(ws, df: pd.DataFrame):
    from openpyxl.utils import get_column_letter
    cols = list(df.columns) if df is not None else []
    for i, col in enumerate(cols, start=1):
        if df is not None and not df.empty:
            sample = [len(str(x)) for x in df[col].head(1000).tolist()]
            max_len = max([len(str(col))] + sample)
        else:
            max_len = max(len(str(col)), 8)
        ws.column_dimensions[get_column_letter(i)].width = min(max_len + 2, 60)

def break_minutes_for_span(first_dt: pd.Timestamp, last_dt: pd.Timestamp) -> Tuple[int,str]:
    if pd.isna(first_dt) or pd.isna(last_dt):
        return 0, "無時間資料"
    st, ed = first_dt.time(), last_dt.time()
    for st_ge, ed_le, mins, tag in BREAK_RULES:
        if (st >= st_ge) and (ed <= ed_le):
            return mins, tag
    return 0, "未命中規則"

def _subtract_exclusions(s_dt: pd.Timestamp, e_dt: pd.Timestamp, exclude_ranges):
    if s_dt >= e_dt or not exclude_ranges:
        return [(s_dt, e_dt)]
    segments = [(s_dt, e_dt)]
    for ex_s_t, ex_e_t in exclude_ranges:
        ex_s = pd.Timestamp.combine(s_dt.date(), ex_s_t)
        ex_e = pd.Timestamp.combine(s_dt.date(), ex_e_t)
        new_segments = []
        for a, b in segments:
            if b <= ex_s or a >= ex_e:
                new_segments.append((a, b))
            else:
                if a < ex_s:
                    new_segments.append((a, ex_s))
                if b > ex_e:
                    new_segments.append((ex_e, b))
        segments = [(x, y) for (x, y) in new_segments if x < y]
    return segments

def _compute_idle(series_dt: pd.Series, min_minutes: int, exclude_ranges) -> Tuple[int, str]:
    if series_dt.size < 2:
        return 0, ""
    s = series_dt.sort_values()
    total_min, ranges_txt = 0, []
    prev = s.iloc[0]
    for cur in s.iloc[1:]:
        if cur <= prev:
            prev = cur
            continue
        for a, b in _subtract_exclusions(prev, cur, exclude_ranges or []):
            gap_min = int(round((b - a).total_seconds() / 60.0))
            if gap_min >= min_minutes:
                total_min += gap_min
                ranges_txt.append(f"{a.time()} ~ {b.time()}")
        prev = cur
    return int(total_min), "；".join(ranges_txt)

def _span_metrics(series_dt: pd.Series):
    if series_dt.empty:
        return pd.NaT, pd.NaT, 0
    return series_dt.min(), series_dt.max(), series_dt.size

def compute_am_pm_for_group(g: pd.DataFrame, *, idle_threshold: int) -> pd.Series:
    times = g["__dt__"]

    t_am = times[times.dt.time.between(AM_START, AM_END)]
    am_first, am_last, am_cnt = _span_metrics(t_am)
    am_mins = int(round(((am_last - am_first).total_seconds()/60.0))) if am_cnt > 0 else 0
    am_eff  = round((am_cnt / am_mins * 60.0), 2) if am_mins > 0 else 0.0
    am_idle_min, am_idle_ranges = _compute_idle(t_am, idle_threshold, EXCLUDE_IDLE_RANGES)

    t_pm = times[times.dt.time.between(PM_START, PM_END)]
    pm_first, pm_last, pm_cnt = _span_metrics(t_pm)
    if pm_cnt > 0:
        pm_break, pm_rule = break_minutes_for_span(pm_first, pm_last)
        raw_pm_mins = (pm_last - pm_first).total_seconds()/60.0
        pm_mins = max(int(round(raw_pm_mins - pm_break)), 0)
    else:
        pm_break, pm_rule, pm_mins = 0, "無時間資料", 0
    pm_eff = round((pm_cnt / pm_mins * 60.0), 2) if pm_mins > 0 else 0.0
    pm_idle_min, pm_idle_ranges = _compute_idle(t_pm, idle_threshold, EXCLUDE_IDLE_RANGES)

    whole_first, whole_last, day_cnt = _span_metrics(times)
    if day_cnt > 0:
        whole_break, br_tag_whole = break_minutes_for_span(whole_first, whole_last)
        raw_whole_mins = (whole_last - whole_first).total_seconds()/60.0
        whole_mins = max(int(round(raw_whole_mins - whole_break)), 0)
    else:
        whole_break, br_tag_whole, whole_mins = 0, "無時間資料", 0
    whole_eff = round((day_cnt / whole_mins * 60.0), 2) if whole_mins > 0 else 0.0

    return pd.Series({
        "第一筆時間": whole_first, "最後一筆時間": whole_last, "當日筆數": int(day_cnt),
        "休息分鐘_整體": int(whole_break), "命中規則": br_tag_whole,
        "當日工時_分鐘_扣休": int(whole_mins), "效率_件每小時": whole_eff,
        "上午_第一筆": am_first, "上午_最後一筆": am_last, "上午_筆數": int(am_cnt),
        "上午_工時_分鐘": int(am_mins), "上午_效率_件每小時": am_eff,
        "上午_空窗分鐘": int(am_idle_min), "上午_空窗時段": am_idle_ranges,
        "下午_第一筆": pm_first, "下午_最後一筆": pm_last, "下午_筆數": int(pm_cnt),
        "下午_休息分鐘": int(pm_break), "下午_命中規則": pm_rule,
        "下午_工時_分鐘_扣休": int(pm_mins), "下午_效率_件每小時": pm_eff,
        "下午_空窗分鐘_扣休": int(pm_idle_min), "下午_空窗時段": pm_idle_ranges,
    })

def shade_rows_by_efficiency(ws, header_name="效率_件每小時", target_eff: float = 20.0, green="C6EFCE", red="FFC7CE"):
    from openpyxl.styles import PatternFill
    eff_col = None
    for c in range(1, ws.max_column + 1):
        if str(ws.cell(row=1, column=c).value).strip() == header_name:
            eff_col = c; break
    if eff_col is None: 
        return
    green_fill = PatternFill(start_color=green, end_color=green, fill_type="solid")
    red_fill   = PatternFill(start_color=red,   end_color=red,   fill_type="solid")
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=eff_col).value
        try:
            val = float(v) if v is not None and str(v).strip() != "" else None
        except Exception:
            val = None
        if val is None: 
            continue
        fill = green_fill if val >= target_eff else red_fill
        for c in range(1, ws.max_column + 1):
            ws.cell(row=r, column=c).fill = fill

def write_block_report(writer, detail_long: pd.DataFrame, user_col: str, target_eff: float):
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    sheet_name = "報表_區塊"
    wb = writer.book
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    header = ["代碼","姓名","筆數","工作區間","總分鐘","效率(件/時)","休息分鐘","空窗分鐘","空窗時段"]
    title_font = Font(bold=True, size=14)
    sec_font   = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    border = Border(left=Side(style="thin"), right=Side(style="thin"),
                    top=Side(style="thin"), bottom=Side(style="thin"))
    center = Alignment(horizontal="center", vertical="center")
    left   = Alignment(horizontal="left",   vertical="center")

    df = detail_long.copy()
    df["工作區間"] = df.apply(lambda r: (
        ("" if pd.isna(r["第一筆時間"]) else str(r["第一筆時間"].time())) + " ~ " +
        ("" if pd.isna(r["最後一筆時間"]) else str(r["最後一筆時間"].time()))
    ), axis=1)
    df["總分鐘"] = df["工時_分鐘"].astype(int)

    for dt_date, g in df.groupby("日期"):
        row = ws.max_row + 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(header))
        cell = ws.cell(row=row, column=1, value=f"{dt_date} 上架績效")
        cell.font = title_font; cell.alignment = center

        for seg in ["上午", "下午"]:
            seg_df = g[g["時段"] == seg]
            if seg_df.empty: 
                continue
            row = ws.max_row + 1
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(header))
            cell = ws.cell(row=row, column=1, value=seg)
            cell.font = sec_font; cell.alignment = left

            row = ws.max_row + 1
            for c, h in enumerate(header, start=1):
                cell = ws.cell(row=row, column=c, value=h)
                cell.fill = header_fill; cell.alignment = center; cell.border = border; cell.font = Font(bold=True)

            seg_df = seg_df.sort_values(["效率_件每小時","筆數"], ascending=[False, False])
            for _, r in seg_df.iterrows():
                row = ws.max_row + 1
                values = [
                    r[user_col], r["對應姓名"], int(r["筆數"]), r["工作區間"],
                    int(r["總分鐘"]), float(r["效率_件每小時"]), int(r["休息分鐘"]),
                    int(r["空窗分鐘"]), r["空窗時段"]
                ]
                for c, v in enumerate(values, start=1):
                    cell = ws.cell(row=row, column=c, value=v)
                    cell.alignment = center if c not in (4,9) else left
                    cell.border = border
                eff = float(r["效率_件每小時"]) if pd.notna(r["效率_件每小時"]) else 0.0
                color = "C6EFCE" if eff >= target_eff else "FFC7CE"
                fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                for c in range(1, len(header)+1):
                    ws.cell(row=row, column=c).fill = fill

    from openpyxl.utils import get_column_letter
    for c in range(1, len(header)+1):
        max_len = max(len(str(header[c-1])),
                      max((len(str(ws.cell(row=r, column=c).value)) for r in range(1, ws.max_row+1)), default=0))
        ws.column_dimensions[get_column_letter(c)].width = min(max_len + 2, 60)

def run_shelf_efficiency(file_bytes: bytes, filename: str, params: Dict[str, Any] | None = None) -> Dict[str, Any]:
    params = params or {}
    target_eff = float(params.get("target_eff", DEFAULT_TARGET_EFF))
    idle_threshold = int(params.get("idle_threshold", DEFAULT_IDLE_MIN_THRESHOLD))

    suffix = os.path.splitext(filename)[1].lower() or ".xlsx"

    with tempfile.TemporaryDirectory() as td:
        in_path = os.path.join(td, f"upload{suffix}")
        with open(in_path, "wb") as f:
            f.write(file_bytes)

        sheets = read_excel_any_quiet(in_path)

        kept_all = []
        for sn, df in sheets.items():
            k = prepare_filtered_df(df)
            if not k.empty:
                k["__sheet__"] = sn
                kept_all.append(k)

        if not kept_all:
            raise Exception("無符合資料（可能缺『由/到』欄或過濾後為空）。")

        data = pd.concat(kept_all, ignore_index=True)

        user_col = find_first_column(data, INPUT_USER_CANDIDATES)
        revdt_col = find_first_column(data, REV_DT_CANDIDATES)
        if user_col is None:
            raise Exception("找不到『記錄輸入人』欄位。")
        if revdt_col is None:
            raise Exception("找不到『修訂日期/時間』欄位。")

        data["__dt__"] = pd.to_datetime(data[revdt_col], errors="coerce")
        data["__code__"] = data[user_col].astype(str).str.strip()
        data["對應姓名"] = data["__code__"].map(NAME_MAP).fillna("")

        dt_data = data.dropna(subset=["__dt__"]).copy()
        if dt_data.empty:
            raise Exception("資料沒有可用的修訂日期時間，無法計算。")

        dt_data["日期"] = dt_data["__dt__"].dt.date

        daily = (
            dt_data.groupby([user_col, "對應姓名", "日期"], dropna=False)
                   .apply(lambda g: compute_am_pm_for_group(g, idle_threshold=idle_threshold))
                   .reset_index()
        )

        # 彙總
        summary = (
            daily.groupby([user_col, "對應姓名"], dropna=False, as_index=False)
                 .agg(
                     総日數=("日期", "nunique"),
                     總筆數=("當日筆數", "sum"),
                     總工時_分鐘_扣休=("當日工時_分鐘_扣休", "sum"),
                     上午筆數=("上午_筆數", "sum"),
                     上午工時_分鐘=("上午_工時_分鐘", "sum"),
                     下午筆數=("下午_筆數", "sum"),
                     下午工時_分鐘_扣休=("下午_工時_分鐘_扣休", "sum"),
                 )
        )

        def _eff(n, m):
            return round((n / m * 60.0), 2) if m and m > 0 else 0.0

        summary["上午效率_件每小時"] = summary.apply(lambda r: _eff(r["上午筆數"], r["上午工時_分鐘"]), axis=1)
        summary["下午效率_件每小時"] = summary.apply(lambda r: _eff(r["下午筆數"], r["下午工時_分鐘_扣休"]), axis=1)
        summary["總工時_分鐘_扣休"] = summary["上午工時_分鐘"].fillna(0).astype(int) + summary["下午工時_分鐘_扣休"].fillna(0).astype(int)
        summary["效率_件每小時"] = summary.apply(lambda r: _eff(r["總筆數"], r["總工時_分鐘_扣休"]), axis=1)

        for c in ["總筆數","總工時_分鐘_扣休","上午筆數","上午工時_分鐘","下午筆數","下午工時_分鐘_扣休"]:
            summary[c] = summary[c].fillna(0).astype(int)
        summary = summary.sort_values(["總筆數","總工時_分鐘_扣休"], ascending=[False, False])

        total_people = int(summary[user_col].nunique())
        met_people = int((summary["效率_件每小時"] >= target_eff).sum())
        rate = (met_people / total_people) if total_people > 0 else 0.0

        total_row = {
            user_col: "整體合計", "對應姓名": "",
            "総日數": int(summary["総日數"].sum()),
            "總筆數": int(summary["總筆數"].sum()),
            "總工時_分鐘_扣休": int(summary["總工時_分鐘_扣休"].sum()),
            "上午筆數": int(summary["上午筆數"].sum()),
            "上午工時_分鐘": int(summary["上午工時_分鐘"].sum()),
            "下午筆數": int(summary["下午筆數"].sum()),
            "下午工時_分鐘_扣休": int(summary["下午工時_分鐘_扣休"].sum()),
            "效率_件每小時": _eff(int(summary["總筆數"].sum()), int(summary["總工時_分鐘_扣休"].sum())),
            "上午效率_件每小時": _eff(int(summary["上午筆數"].sum()), int(summary["上午工時_分鐘"].sum())),
            "下午效率_件每小時": _eff(int(summary["下午筆數"].sum()), int(summary["下午工時_分鐘_扣休"].sum())),
        }
        summary_out = pd.concat([summary, pd.DataFrame([total_row])], ignore_index=True)

        # 明細_時段（長表）
        long_rows = []
        for _, r in daily.iterrows():
            if r["上午_筆數"] > 0:
                long_rows.append({
                    user_col: r[user_col], "對應姓名": r["對應姓名"], "日期": r["日期"],
                    "時段": "上午",
                    "第一筆時間": r["上午_第一筆"], "最後一筆時間": r["上午_最後一筆"],
                    "筆數": int(r["上午_筆數"]),
                    "工時_分鐘": int(r["上午_工時_分鐘"]),
                    "休息分鐘": 0,
                    "空窗分鐘": int(r["上午_空窗分鐘"]),
                    "空窗時段": r["上午_空窗時段"],
                    "效率_件每小時": r["上午_效率_件每小時"],
                    "命中規則": "上午不扣休",
                })
            if r["下午_筆數"] > 0:
                long_rows.append({
                    user_col: r[user_col], "對應姓名": r["對應姓名"], "日期": r["日期"],
                    "時段": "下午",
                    "第一筆時間": r["下午_第一筆"], "最後一筆時間": r["下午_最後一筆"],
                    "筆數": int(r["下午_筆數"]),
                    "工時_分鐘": int(r["下午_工時_分鐘_扣休"]),
                    "休息分鐘": int(r["下午_休息分鐘"]),
                    "空窗分鐘": int(r["下午_空窗分鐘_扣休"]),
                    "空窗時段": r["下午_空窗時段"],
                    "效率_件每小時": r["下午_效率_件每小時"],
                    "命中規則": r["下午_命中規則"],
                })
        detail_long = pd.DataFrame(long_rows)
        if not detail_long.empty:
            detail_long = detail_long.sort_values([user_col,"日期","時段","第一筆時間"])

        # 匯出 Excel（保留著色與報表）
        base = os.path.splitext(os.path.basename(filename))[0]
        xlsx_name = f"{base}上架績效.xlsx"
        out_path = os.path.join(td, "result.xlsx")

        with pd.ExcelWriter(out_path, engine="openpyxl",
                            datetime_format="yyyy-mm-dd hh:mm:ss",
                            date_format="yyyy-mm-dd") as writer:
            sum_cols = [
                user_col, "對應姓名", "総日數",
                "總筆數","總工時_分鐘_扣休","效率_件每小時",
                "上午筆數","上午工時_分鐘","上午效率_件每小時",
                "下午筆數","下午工時_分鐘_扣休","下午效率_件每小時",
            ]
            summary_out[sum_cols].to_excel(writer, index=False, sheet_name="彙總")
            ws_sum = writer.sheets["彙總"]; autosize_columns(ws_sum, summary_out[sum_cols])

            det_cols = [
                user_col, "對應姓名", "日期",
                "第一筆時間","最後一筆時間","當日筆數",
                "休息分鐘_整體","當日工時_分鐘_扣休","效率_件每小時",
                "上午_第一筆","上午_最後一筆","上午_筆數","上午_工時_分鐘","上午_效率_件每小時",
                "上午_空窗分鐘","上午_空窗時段",
                "下午_第一筆","下午_最後一筆","下午_筆數","下午_休息分鐘",
                "下午_工時_分鐘_扣休","下午_效率_件每小時",
                "下午_空窗分鐘_扣休","下午_空窗時段",
            ]
            daily.sort_values([user_col,"日期","第一筆時間"])[det_cols].to_excel(writer, index=False, sheet_name="明細")
            ws_det = writer.sheets["明細"]; autosize_columns(ws_det, daily[det_cols])

            if not detail_long.empty:
                long_cols = [user_col,"對應姓名","日期","時段","第一筆時間","最後一筆時間",
                             "筆數","工時_分鐘","休息分鐘","空窗分鐘","空窗時段",
                             "效率_件每小時","命中規則"]
                detail_long[long_cols].to_excel(writer, index=False, sheet_name="明細_時段")
                ws_long = writer.sheets["明細_時段"]; autosize_columns(ws_long, detail_long[long_cols])
                shade_rows_by_efficiency(ws_long, header_name="效率_件每小時", target_eff=target_eff)

                write_block_report(writer, detail_long, user_col, target_eff=target_eff)

            rules_rows = []
            for i,(st_ge,ed_le,mins,tag) in enumerate(BREAK_RULES, start=1):
                rules_rows.append({
                    "優先序": i,
                    "首時間條件(>=)": st_ge.strftime("%H:%M:%S"),
                    "末時間條件(<=)": ed_le.strftime("%H:%M:%S"),
                    "休息分鐘": mins,
                    "規則說明": tag
                })
            rules_df = pd.DataFrame(rules_rows, columns=["優先序","首時間條件(>=)","末時間條件(<=)","休息分鐘","規則說明"])
            rules_df.to_excel(writer, index=False, sheet_name="休息規則")
            ws_rule = writer.sheets["休息規則"]; autosize_columns(ws_rule, rules_df)

            shade_rows_by_efficiency(ws_sum, header_name="效率_件每小時", target_eff=target_eff)
            shade_rows_by_efficiency(ws_det, header_name="效率_件每小時", target_eff=target_eff)

        with open(out_path, "rb") as f:
            xlsx_bytes = f.read()

    # UI 用的彙總欄位（統一名稱方便共用 UI）
    ui_summary = summary_out.copy()
    ui_summary = ui_summary.rename(columns={
        user_col: "記錄輸入人",
        "對應姓名": "姓名",
        "總筆數": "筆數",
        "總工時_分鐘_扣休": "總分鐘",
        "效率_件每小時": "效率",
    })

    # 轉小時給 KPI 顯示
    total_minutes = int(summary_out.loc[summary_out.index[:-1], "總工時_分鐘_扣休"].sum()) if len(summary_out)>1 else int(summary_out["總工時_分鐘_扣休"].sum())
    total_hours = round(total_minutes / 60.0, 2) if total_minutes else 0.0
    avg_eff = round(float(summary["效率_件每小時"].mean()), 2) if len(summary)>0 else 0.0

    return {
        "summary_df": ui_summary,
        "detail_df": daily,
        "ampm_df": detail_long.rename(columns={user_col: "記錄輸入人", "對應姓名":"姓名"}) if not detail_long.empty else pd.DataFrame(),
        "xlsx_bytes": xlsx_bytes,
        "xlsx_name": xlsx_name,
        "target_eff": target_eff,
        "people": total_people,
        "total_count": int(summary["總筆數"].sum()) if len(summary)>0 else 0,
        "total_hours": total_hours,
        "avg_eff": avg_eff,
        "pass_rate": f"{rate:.0%}",
    }
